import base64
import hashlib
import io
import json
import math
import secrets
import ssl
import tempfile
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, List, Tuple

import ezdxf
from PIL import Image, ImageDraw
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import streamlit.components.v1 as components
import streamlit.elements.image as st_image
from streamlit.elements.lib.image_utils import image_to_url as _streamlit_image_to_url
from streamlit.elements.lib.layout_utils import LayoutConfig
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from streamlit_plotly_events import plotly_events
from streamlit_drawable_canvas import st_canvas

try:
    import certifi
except ModuleNotFoundError:
    certifi = None

try:
    import tomllib
except ModuleNotFoundError:
    tomllib = None

st.set_page_config(page_title="Dimensionamento Eletrico Residencial", layout="wide")

if not hasattr(st_image, "image_to_url"):
    def _compat_image_to_url(image, width, clamp, channels, output_format, image_id):
        return _streamlit_image_to_url(
            image=image,
            layout_config=LayoutConfig(width=width),
            clamp=clamp,
            channels=channels,
            output_format=output_format,
            image_id=image_id,
        )

    st_image.image_to_url = _compat_image_to_url


def _session_key(prefix: str) -> str:
    return f"dxf_import_{prefix}"


def _hash_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _extract_segments(doc: ezdxf.EzDxfDocument) -> list[tuple[tuple[float, float], tuple[float, float]]]:
    msp = doc.modelspace()
    segments: list[tuple[tuple[float, float], tuple[float, float]]] = []

    for entity in msp.query("LINE"):
        start = (float(entity.dxf.start.x), float(entity.dxf.start.y))
        end = (float(entity.dxf.end.x), float(entity.dxf.end.y))
        if start != end:
            segments.append((start, end))

    for entity in msp.query("LWPOLYLINE"):
        points = [(float(x), float(y)) for x, y, *_ in entity.get_points("xy")]
        if entity.closed and points and points[0] != points[-1]:
            points.append(points[0])
        for i in range(len(points) - 1):
            if points[i] != points[i + 1]:
                segments.append((points[i], points[i + 1]))

    for entity in msp.query("POLYLINE"):
        if entity.get_mode() != "AcDb2dPolyline":
            continue
        points = [(float(vertex.dxf.location.x), float(vertex.dxf.location.y)) for vertex in entity.vertices]
        if entity.is_closed and points and points[0] != points[-1]:
            points.append(points[0])
        for i in range(len(points) - 1):
            if points[i] != points[i + 1]:
                segments.append((points[i], points[i + 1]))

    return segments


def _load_dxf_payload(file_bytes: bytes) -> tuple[list[tuple[tuple[float, float], tuple[float, float]]], list[tuple[float, float]]]:
    tmp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".dxf") as tmp:
            tmp.write(file_bytes)
            tmp_path = Path(tmp.name)
        doc = ezdxf.readfile(tmp_path)
        segments = _extract_segments(doc)
        if not segments:
            raise ValueError("Nenhuma entidade grafica utilizavel foi encontrada no DXF.")
        points = sorted({point for segment in segments for point in segment})
        return segments, points
    finally:
        if tmp_path and tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass


def _polygon_area(points: list[tuple[float, float]]) -> float:
    acc = 0.0
    total = len(points)
    for i in range(total):
        x1, y1 = points[i]
        x2, y2 = points[(i + 1) % total]
        acc += x1 * y2 - x2 * y1
    return abs(acc) / 2.0


def _polygon_perimeter(points: list[tuple[float, float]]) -> float:
    total = 0.0
    count = len(points)
    for i in range(count):
        x1, y1 = points[i]
        x2, y2 = points[(i + 1) % count]
        total += math.hypot(x2 - x1, y2 - y1)
    return total


def _polygon_centroid(points: list[tuple[float, float]]) -> tuple[float, float]:
    area_factor = 0.0
    cx = 0.0
    cy = 0.0
    count = len(points)
    for i in range(count):
        x1, y1 = points[i]
        x2, y2 = points[(i + 1) % count]
        cross = x1 * y2 - x2 * y1
        area_factor += cross
        cx += (x1 + x2) * cross
        cy += (y1 + y2) * cross

    if abs(area_factor) < 1e-9:
        xs = [point[0] for point in points]
        ys = [point[1] for point in points]
        return sum(xs) / len(xs), sum(ys) / len(ys)

    area_factor *= 0.5
    cx /= 6.0 * area_factor
    cy /= 6.0 * area_factor
    return cx, cy


def _distance(a: tuple[float, float], b: tuple[float, float]) -> float:
    return math.hypot(a[0] - b[0], a[1] - b[1])



def _normalize_room_geometry(
    room: Dict[str, Any],
    endpoints: list[tuple[float, float]],
) -> Dict[str, Any]:
    room_norm = dict(room)
    vertices = room_norm.get('vertices') or []
    if not endpoints:
        return room_norm

    min_x, max_x, min_y, max_y, pad_x, pad_y = _project_bounds(endpoints)

    def centroid_of(points: list[tuple[float, float]]) -> tuple[float, float]:
        if not points:
            return 0.0, 0.0
        return _polygon_centroid(points)

    def is_inside_project(points: list[tuple[float, float]]) -> bool:
        if not points:
            return False
        cx, cy = centroid_of(points)
        return (min_x - pad_x) <= cx <= (max_x + pad_x) and (min_y - pad_y) <= cy <= (max_y + pad_y)

    if vertices and is_inside_project(vertices):
        return room_norm

    capture_transform = room_norm.get('capture_transform') or {}
    canvas_vertices = room_norm.get('canvas_vertices') or []
    if canvas_vertices and capture_transform:
        recovered = [_canvas_to_world_point(point, capture_transform) for point in canvas_vertices]
        if recovered and is_inside_project(recovered):
            room_norm['vertices'] = recovered
            cx, cy = _polygon_centroid(recovered)
            room_norm['centroide_x'] = cx
            room_norm['centroide_y'] = cy
            return room_norm

    if vertices:
        cx, cy = centroid_of(vertices)
        room_norm['centroide_x'] = cx
        room_norm['centroide_y'] = cy
    return room_norm


def _nearest_endpoint(clicked: tuple[float, float], endpoints: list[tuple[float, float]]) -> tuple[float, float]:
    return min(endpoints, key=lambda point: _distance(point, clicked))


def _constrain_to_orthogonal(
    clicked: tuple[float, float],
    last_point: tuple[float, float],
) -> tuple[float, float]:
    dx = clicked[0] - last_point[0]
    dy = clicked[1] - last_point[1]
    if abs(dx) >= abs(dy):
        return clicked[0], last_point[1]
    return last_point[0], clicked[1]


DXF_MARGIN_FACTOR = 0.03


def _apply_dxf_margin(valor: float) -> float:
    return math.ceil((valor * (1.0 + DXF_MARGIN_FACTOR)) * 10.0) / 10.0


def _infer_tipo_comodo(nome: str) -> str:
    nome_normalizado = normalizar_ambiente(nome)
    nome_bruto = nome.strip().lower()

    if 'lavabo' in nome_bruto:
        return 'lavabo'
    if 'banheiro' in nome_bruto or 'banho' in nome_bruto or nome_normalizado == 'banheiro':
        return 'banheiro'
    if 'suite' in nome_bruto or nome_normalizado == 'suite':
        return 'suite'
    if 'dorm' in nome_bruto or 'quarto' in nome_bruto or nome_normalizado == 'quarto':
        return 'quarto'
    if 'cozinha' in nome_bruto or nome_normalizado == 'cozinha':
        return 'cozinha'
    if 'serv' in nome_bruto:
        return 'area_servico'
    if 'varanda' in nome_bruto:
        return 'varanda'
    if 'corredor' in nome_bruto:
        return 'corredor'
    if 'circ' in nome_bruto or nome_normalizado == 'circulacao':
        return 'circulacao'
    if 'hall' in nome_bruto:
        return 'hall'
    if 'jantar' in nome_bruto or 'estar' in nome_bruto or nome_normalizado == 'sala':
        return 'sala'
    return nome_normalizado if nome_normalizado else 'outro'


def _build_imported_room(points: list[tuple[float, float]], index: int) -> dict[str, Any]:
    area_bruta = _polygon_area(points)
    perimetro_bruto = _polygon_perimeter(points)
    centroide_x, centroide_y = _polygon_centroid(points)
    area = _apply_dxf_margin(area_bruta)
    perimetro = _apply_dxf_margin(perimetro_bruto)
    return {
        "nome": f"Comodo {index}",
        "tipo_sugerido": 'outro',
        "area": area,
        "perimetro": perimetro,
        "area_bruta": area_bruta,
        "perimetro_bruto": perimetro_bruto,
        "centroide_x": centroide_x,
        "centroide_y": centroide_y,
        "vertices": points[:],
    }


def _project_bounds(
    endpoints: list[tuple[float, float]],
) -> tuple[float, float, float, float, float, float]:
    xs = [point[0] for point in endpoints]
    ys = [point[1] for point in endpoints]
    min_x, max_x = min(xs), max(xs)
    min_y, max_y = min(ys), max(ys)
    pad_x = max((max_x - min_x) * 0.04, 1.0)
    pad_y = max((max_y - min_y) * 0.04, 1.0)
    return min_x, max_x, min_y, max_y, pad_x, pad_y


def _default_view_state(endpoints: list[tuple[float, float]]) -> dict[str, float]:
    min_x, max_x, min_y, max_y, pad_x, pad_y = _project_bounds(endpoints)
    return {
        'center_x': (min_x + max_x) / 2.0,
        'center_y': (min_y + max_y) / 2.0,
        'zoom': 1.0,
        'min_x': min_x,
        'max_x': max_x,
        'min_y': min_y,
        'max_y': max_y,
        'pad_x': pad_x,
        'pad_y': pad_y,
    }


def _clamp_view_state(view_state: dict[str, float]) -> dict[str, float]:
    min_x = view_state['min_x'] - view_state['pad_x']
    max_x = view_state['max_x'] + view_state['pad_x']
    min_y = view_state['min_y'] - view_state['pad_y']
    max_y = view_state['max_y'] + view_state['pad_y']
    full_width = max(max_x - min_x, 1.0)
    full_height = max(max_y - min_y, 1.0)

    zoom = min(max(view_state.get('zoom', 1.0), 1.0), 8.0)
    visible_width = full_width / zoom
    visible_height = full_height / zoom

    half_w = visible_width / 2.0
    half_h = visible_height / 2.0

    center_x = min(max(view_state.get('center_x', (min_x + max_x) / 2.0), min_x + half_w), max_x - half_w)
    center_y = min(max(view_state.get('center_y', (min_y + max_y) / 2.0), min_y + half_h), max_y - half_h)

    return {
        **view_state,
        'center_x': center_x,
        'center_y': center_y,
        'zoom': zoom,
    }


def _update_view_state(view_state: dict[str, float], action: str) -> dict[str, float]:
    state = _clamp_view_state(view_state)
    min_x = state['min_x'] - state['pad_x']
    max_x = state['max_x'] + state['pad_x']
    min_y = state['min_y'] - state['pad_y']
    max_y = state['max_y'] + state['pad_y']
    full_width = max(max_x - min_x, 1.0)
    full_height = max(max_y - min_y, 1.0)
    visible_width = full_width / state['zoom']
    visible_height = full_height / state['zoom']
    move_x = visible_width * 0.12
    move_y = visible_height * 0.12

    updated = dict(state)
    if action == 'zoom_in':
        updated['zoom'] = min(state['zoom'] * 1.25, 8.0)
    elif action == 'zoom_out':
        updated['zoom'] = max(state['zoom'] / 1.25, 1.0)
    elif action == 'left':
        updated['center_x'] = state['center_x'] - move_x
    elif action == 'right':
        updated['center_x'] = state['center_x'] + move_x
    elif action == 'up':
        updated['center_y'] = state['center_y'] + move_y
    elif action == 'down':
        updated['center_y'] = state['center_y'] - move_y
    elif action == 'reset':
        updated['center_x'] = (state['min_x'] + state['max_x']) / 2.0
        updated['center_y'] = (state['min_y'] + state['max_y']) / 2.0
        updated['zoom'] = 1.0
    return _clamp_view_state(updated)


def _build_canvas_background(
    segments: list[tuple[tuple[float, float], tuple[float, float]]],
    endpoints: list[tuple[float, float]],
    view_state: dict[str, float],
    max_width: int = 1500,
    max_height: int = 980,
) -> tuple[Image.Image, dict[str, float]]:
    state = _clamp_view_state(view_state)
    min_x = state['min_x'] - state['pad_x']
    max_x = state['max_x'] + state['pad_x']
    min_y = state['min_y'] - state['pad_y']
    max_y = state['max_y'] + state['pad_y']
    full_width = max(max_x - min_x, 1.0)
    full_height = max(max_y - min_y, 1.0)

    visible_width = full_width / state['zoom']
    visible_height = full_height / state['zoom']
    left = state['center_x'] - visible_width / 2.0
    right = state['center_x'] + visible_width / 2.0
    bottom = state['center_y'] - visible_height / 2.0
    top = state['center_y'] + visible_height / 2.0

    canvas_width = max_width
    canvas_height = max_height
    scale = min(canvas_width / visible_width, canvas_height / visible_height)
    draw_width = visible_width * scale
    draw_height = visible_height * scale
    offset_x = (canvas_width - draw_width) / 2.0
    offset_y = (canvas_height - draw_height) / 2.0

    image = Image.new('RGB', (canvas_width, canvas_height), '#1f2937')
    draw = ImageDraw.Draw(image)

    def to_canvas(point: tuple[float, float]) -> tuple[float, float]:
        x = (point[0] - left) * scale + offset_x
        y = (top - point[1]) * scale + offset_y
        return x, y

    for start_point, end_point in segments:
        x1, y1 = to_canvas(start_point)
        x2, y2 = to_canvas(end_point)
        draw.line((x1, y1, x2, y2), fill='#e5e7eb', width=2)

    transform = {
        'left': left,
        'top': top,
        'scale': scale,
        'offset_x': offset_x,
        'offset_y': offset_y,
        'canvas_width': float(canvas_width),
        'canvas_height': float(canvas_height),
    }
    return image, transform


def _canvas_to_world_point(point: tuple[float, float], transform: dict[str, float]) -> tuple[float, float]:
    x = (point[0] - transform['offset_x']) / transform['scale'] + transform['left']
    y = transform['top'] - ((point[1] - transform['offset_y']) / transform['scale'])
    return round(x, 4), round(y, 4)


def _world_to_canvas_point(point: tuple[float, float], transform: dict[str, float]) -> tuple[float, float]:
    x = (point[0] - transform['left']) * transform['scale'] + transform['offset_x']
    y = (transform['top'] - point[1]) * transform['scale'] + transform['offset_y']
    return round(x, 2), round(y, 2)


def _extract_polygon_points_from_canvas_object(obj: dict[str, Any]) -> list[tuple[float, float]]:
    if obj.get('type') == 'path' and obj.get('path'):
        points: list[tuple[float, float]] = []
        for command in obj.get('path', []):
            if not command:
                continue
            opcode = command[0]
            if opcode in ('M', 'L') and len(command) >= 3:
                x = float(command[1])
                y = float(command[2])
                point = (x, y)
                if not points or point != points[-1]:
                    points.append(point)
        if len(points) > 1 and points[0] == points[-1]:
            points.pop()
        return points

    if obj.get('type') == 'polygon' and obj.get('points'):
        left = float(obj.get('left', 0.0))
        top = float(obj.get('top', 0.0))
        path_offset = obj.get('pathOffset') or {}
        offset_x = float(path_offset.get('x', 0.0))
        offset_y = float(path_offset.get('y', 0.0))
        points = []
        for point in obj.get('points', []):
            x = float(point.get('x', 0.0)) + left - offset_x
            y = float(point.get('y', 0.0)) + top - offset_y
            points.append((x, y))
        return points

    return []


def _extract_latest_room_from_canvas(
    canvas_json: dict[str, Any] | None,
    transform: dict[str, float],
    index: int,
) -> dict[str, Any] | None:
    if not canvas_json:
        return None
    objects = canvas_json.get('objects') or []
    for obj in reversed(objects):
        points_canvas = _extract_polygon_points_from_canvas_object(obj)
        if len(points_canvas) >= 3:
            points_world = [_canvas_to_world_point(point, transform) for point in points_canvas]
            room = _build_imported_room(points_world, index)
            room['canvas_vertices'] = points_canvas[:]
            room['capture_transform'] = dict(transform)
            return room
    return None


def _inject_dxf_hotkeys(prefix: str) -> None:
    components.html(
        f"""
        <script>
        (() => {{
          const flag = '__dxf_hotkeys_{prefix}__';
          if (window.parent[flag]) return;
          window.parent[flag] = true;

          const findParentButton = (label) => Array.from(window.parent.document.querySelectorAll('button')).find((btn) => (btn.innerText || '').trim() === label);
          const clickToolbarByAlt = (altText) => {{
            const frames = Array.from(window.parent.document.querySelectorAll('iframe'));
            for (const frame of frames) {{
              try {{
                const doc = frame.contentDocument || frame.contentWindow.document;
                if (!doc) continue;
                const img = doc.querySelector(`img[alt="${{altText}}"]`);
                if (img) {{
                  const button = img.closest('button') || img.parentElement;
                  if (button) {{ button.click(); return true; }}
                }}
              }} catch (err) {{}}
            }}
            return false;
          }};

          const handler = (event) => {{
            const tag = (event.target && event.target.tagName) ? event.target.tagName.toLowerCase() : '';
            if (tag === 'input' || tag === 'textarea') return;

            if (event.key === 'Enter') {{
              const btn = findParentButton('Salvar comodo desenhado');
              if (btn) {{ event.preventDefault(); btn.click(); }}
            }}
            if (event.key === 'Escape') {{
              const btn = findParentButton('Limpar desenho atual');
              if (btn) {{ event.preventDefault(); btn.click(); }}
            }}
            if ((event.ctrlKey || event.metaKey) && event.key.toLowerCase() === 'z') {{
              if (clickToolbarByAlt('Undo last operation')) {{ event.preventDefault(); }}
            }}
            if ((event.ctrlKey || event.metaKey) && event.key.toLowerCase() === 'y') {{
              if (clickToolbarByAlt('Redo last operation')) {{ event.preventDefault(); }}
            }}
            if (event.key === '+') {{
              const btn = findParentButton('Zoom +');
              if (btn) {{ event.preventDefault(); btn.click(); }}
            }}
            if (event.key === '-') {{
              const btn = findParentButton('Zoom -');
              if (btn) {{ event.preventDefault(); btn.click(); }}
            }}
          }};

          window.parent.document.addEventListener('keydown', handler, true);
          window.document.addEventListener('keydown', handler, true);
          const frames = Array.from(window.parent.document.querySelectorAll('iframe'));
          for (const frame of frames) {{
            try {{
              const doc = frame.contentDocument || frame.contentWindow.document;
              if (doc) doc.addEventListener('keydown', handler, true);
            }} catch (err) {{}}
          }}
        }})();
        </script>
        """,
        height=0,
    )


def _render_dxf_editor_panel(
    *,
    segments: list[tuple[tuple[float, float], tuple[float, float]]],
    endpoints: list[tuple[float, float]],
    rooms_key: str,
    canvas_key: str,
    open_key: str,
    view_key: str,
    pending_key: str,
    pending_name_key: str,
    dialog_mode: bool,
) -> None:
    rooms = st.session_state[rooms_key]
    canvas_state = st.session_state.get(canvas_key)
    view_state = st.session_state.get(view_key, _default_view_state(endpoints))
    background_image, transform = _build_canvas_background(segments, endpoints, view_state)

    _inject_dxf_hotkeys('canvas')

    st.markdown(
        """
        <style>
        iframe[title="streamlit_drawable_canvas.st_canvas"] {
            width: 100% !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    reset_pending_name_key = _session_key('reset_pending_name')
    if st.session_state.pop(reset_pending_name_key, False):
        st.session_state.pop(pending_name_key, None)
    if pending_name_key not in st.session_state:
        st.session_state[pending_name_key] = ''

    pending_room = st.session_state.get(pending_key)
    if pending_room:
        with st.container(border=True):
            st.markdown('**Nomear comodo**')
            st.write(f"Area: {pending_room['area']:.2f} m2 | Perimetro: {pending_room['perimetro']:.2f} m")
            nome_digitado = st.text_input('Qual o nome do comodo?', key=pending_name_key)
            confirm_col, cancel_col = st.columns(2)
            with confirm_col:
                confirmar_nome = st.button('Confirmar nome do comodo', use_container_width=True, type='primary')
            with cancel_col:
                cancelar_nome = st.button('Cancelar salvamento', use_container_width=True)

            if confirmar_nome:
                nome_limpo = nome_digitado.strip()
                if not nome_limpo:
                    st.warning('Digite um nome para o comodo antes de confirmar.')
                else:
                    pending_room['nome'] = nome_limpo
                    pending_room['tipo_sugerido'] = _infer_tipo_comodo(nome_limpo)
                    rooms.append(pending_room)
                    st.session_state[rooms_key] = rooms
                    st.session_state[pending_key] = None
                    st.session_state[reset_pending_name_key] = True
                    st.session_state[canvas_key] = None
                    st.success(f"{nome_limpo} salvo com sucesso.")
                    st.rerun()

            if cancelar_nome:
                st.session_state[pending_key] = None
                st.session_state[reset_pending_name_key] = True
                st.rerun()

    nav_cols = st.columns([1, 1, 1, 1, 1, 1, 1, 1] if dialog_mode else [1, 1, 1, 1, 1, 1, 1])
    nav_prefix = _session_key('dialog_nav' if dialog_mode else 'inline_nav')
    with nav_cols[0]:
        zoom_in = st.button('Zoom +', key=f'{nav_prefix}_zoom_in', use_container_width=True)
    with nav_cols[1]:
        zoom_out = st.button('Zoom -', key=f'{nav_prefix}_zoom_out', use_container_width=True)
    with nav_cols[2]:
        move_left = st.button('Esq', key=f'{nav_prefix}_left', use_container_width=True)
    with nav_cols[3]:
        move_right = st.button('Dir', key=f'{nav_prefix}_right', use_container_width=True)
    with nav_cols[4]:
        move_up = st.button('Cima', key=f'{nav_prefix}_up', use_container_width=True)
    with nav_cols[5]:
        move_down = st.button('Baixo', key=f'{nav_prefix}_down', use_container_width=True)
    with nav_cols[6]:
        reset_view = st.button('Centralizar', key=f'{nav_prefix}_reset', use_container_width=True)
    fechar_editor = False
    if dialog_mode:
        with nav_cols[7]:
            fechar_editor = st.button('Fechar editor ampliado', key=f'{nav_prefix}_close', use_container_width=True)

    action = None
    if zoom_in:
        action = 'zoom_in'
    elif zoom_out:
        action = 'zoom_out'
    elif move_left:
        action = 'left'
    elif move_right:
        action = 'right'
    elif move_up:
        action = 'up'
    elif move_down:
        action = 'down'
    elif reset_view:
        action = 'reset'

    if action:
        st.session_state[view_key] = _update_view_state(view_state, action)
        st.session_state[canvas_key] = None
        st.rerun()

    top_cols = st.columns([1.1, 1.0, 1.1, 1.2] if dialog_mode else [1.1, 1.0, 1.1])
    action_prefix = _session_key('dialog_action' if dialog_mode else 'inline_action')
    with top_cols[0]:
        salvar = st.button('Salvar comodo desenhado', key=f'{action_prefix}_save', use_container_width=True)
    with top_cols[1]:
        limpar = st.button('Limpar desenho atual', key=f'{action_prefix}_clear', use_container_width=True)
    with top_cols[2]:
        redefinir = st.button('Redefinir comodos do DXF', key=f'{action_prefix}_reset_rooms', use_container_width=True)

    if redefinir:
        st.session_state[rooms_key] = []
        st.session_state[canvas_key] = None
        st.session_state[pending_key] = None
        st.session_state[reset_pending_name_key] = True
        st.rerun()

    if limpar:
        st.session_state[canvas_key] = None
        st.session_state[pending_key] = None
        st.session_state[reset_pending_name_key] = True
        st.rerun()

    canvas_result = st_canvas(
        fill_color='rgba(16, 185, 129, 0.22)',
        stroke_width=2,
        stroke_color='#22c55e',
        background_image=background_image,
        update_streamlit=True,
        height=int(transform['canvas_height']),
        width=int(transform['canvas_width']),
        drawing_mode='polygon',
        initial_drawing=canvas_state,
        display_toolbar=True,
        point_display_radius=3,
        key=_session_key('canvas_dialog' if dialog_mode else 'canvas_inline'),
    )

    if canvas_result.json_data is not None:
        st.session_state[canvas_key] = canvas_result.json_data
        canvas_state = canvas_result.json_data

    if salvar and not pending_room:
        novo = _extract_latest_room_from_canvas(canvas_state, transform, len(rooms) + 1)
        if novo is None:
            st.warning('Desenhe um poligono fechado no canvas antes de salvar o comodo. Feche com duplo clique ou clique direito.')
        else:
            st.session_state[pending_key] = novo
            st.session_state[reset_pending_name_key] = True
            st.rerun()

    if fechar_editor:
        st.session_state[open_key] = False
        st.rerun()

    st.caption('Clique nos vertices do comodo sobre a planta. Feche o poligono com duplo clique ou clique direito.')
    st.caption('Atalhos: Enter salva, Esc limpa, Ctrl+Z desfaz, Ctrl+Y refaz, + e - controlam o zoom.')

    info_col1, info_col2 = st.columns([1.5, 1])
    with info_col1:
        st.markdown('**Como usar**')
        st.write('1. Ajuste a vista com Zoom e setas, se precisar.')
        st.write('2. Clique nos cantos do c?modo em ordem.')
        st.write('3. Feche com duplo clique ou botao direito.')
        st.write('4. Clique em Salvar comodo desenhado e informe o nome.')
    with info_col2:
        st.markdown('**Comodos importados**')
        if not rooms:
            st.info('Nenhum comodo salvo ainda.')
        else:
            for idx, room in enumerate(rooms, start=1):
                st.markdown(
                    f"**{idx}. {room['nome']}**  \nArea: {room['area']:.2f} m2  \nPerimetro: {room['perimetro']:.2f} m"
                )


def renderizar_importacao_dxf() -> list[dict[str, Any]]:
    st.subheader('1) Importacao DXF e selecao manual dos comodos')
    st.caption('Envie o DXF e abra o editor para desenhar cada comodo diretamente sobre a planta.')

    rooms_key = _session_key('rooms')
    endpoints_key = _session_key('endpoints')
    segments_key = _session_key('segments')
    open_key = _session_key('open_editor')
    canvas_key = _session_key('canvas_state')
    view_key = _session_key('view_state')
    pending_key = _session_key('pending_room')
    pending_name_key = _session_key('pending_room_name')

    if rooms_key not in st.session_state:
        st.session_state[rooms_key] = []
    if open_key not in st.session_state:
        st.session_state[open_key] = False

    uploaded_file = st.file_uploader(
        'Adicionar arquivo DXF',
        type=['dxf'],
        key=_session_key('upload'),
    )

    if uploaded_file is None:
        return st.session_state[rooms_key]

    file_bytes = uploaded_file.getvalue()
    file_hash = _hash_bytes(file_bytes)

    if st.session_state.get(_session_key('file_hash')) != file_hash:
        try:
            segments, endpoints = _load_dxf_payload(file_bytes)
        except Exception as exc:
            st.error(f'Falha ao ler o DXF: {exc}')
            return st.session_state[rooms_key]

        st.session_state[_session_key('file_hash')] = file_hash
        st.session_state[segments_key] = segments
        st.session_state[endpoints_key] = endpoints
        st.session_state[view_key] = _default_view_state(endpoints)
        st.session_state[rooms_key] = []
        st.session_state[canvas_key] = None
        st.session_state[pending_key] = None
        st.session_state[_session_key('reset_pending_name')] = True
        st.session_state[open_key] = False

    segments = st.session_state.get(segments_key)
    endpoints = st.session_state.get(endpoints_key)
    rooms = st.session_state[rooms_key]

    if not segments or not endpoints:
        st.error('Nao foi possivel montar a visualizacao do DXF.')
        return rooms

    action_col1, action_col2 = st.columns([1.4, 2.2])
    with action_col1:
        if st.button('Abrir editor do projeto em tela ampliada', use_container_width=True):
            st.session_state[open_key] = True
            st.rerun()
    with action_col2:
        st.info('No editor, ajuste a vista, desenhe o c?modo, salve e informe o nome na hora.')

    @st.dialog('Editor DXF', width='large')
    def _editor_dialog() -> None:
        _render_dxf_editor_panel(
            segments=segments,
            endpoints=endpoints,
            rooms_key=rooms_key,
            canvas_key=canvas_key,
            open_key=open_key,
            view_key=view_key,
            pending_key=pending_key,
            pending_name_key=pending_name_key,
            dialog_mode=True,
        )

    if st.session_state.get(open_key):
        _editor_dialog()

    st.markdown('**Comodos importados**')
    if not rooms:
        st.warning('Abra o editor ampliado, desenhe e salve ao menos um comodo no DXF para continuar com o formulario abaixo.')
    else:
        resumo_cols = st.columns(min(3, max(1, len(rooms))))
        for idx, room in enumerate(rooms):
            with resumo_cols[idx % len(resumo_cols)]:
                st.metric(room['nome'], f"{room['area']:.2f} m2", f"Perimetro {room['perimetro']:.2f} m")

    return rooms


CATEGORIAS_DEMANDA = {
    "b": "Chuveiro, torneira eletrica, aquecedor de passagem, ferro eletrico",
    "c": "Boiler / aquecedor central",
    "d": "Secadora, forno eletrico, lava-loucas, micro-ondas",
    "e": "Fogao eletrico",
    "f": "Ar-condicionado tipo janela",
    "g": "Motor / maquina de solda a motor",
    "h": "Equipamento especial",
    "i": "Hidromassagem",
    "x": "Sem categoria GED-13 (considerar FD = 1)",
}

PADRAO_ENTRADA_TABELA_1A = {
    "A1": {
        "tipo_caixa": "II",
        "disjuntor_a": 32,
        "medida_eletroduto": "32 (1)",
    },
    "A2": {
        "tipo_caixa": "II",
        "disjuntor_a": 63,
        "medida_eletroduto": "32 (1)",
    },
    "B1": {
        "tipo_caixa": "II",
        "disjuntor_a": 63,
        "medida_eletroduto": "40 (1 1/4)",
    },
    "B2": {
        "tipo_caixa": "II",
        "disjuntor_a": 80,
        "medida_eletroduto": "40 (1 1/4)",
    },
    "C1": {
        "tipo_caixa": "III",
        "disjuntor_a": 63,
        "medida_eletroduto": "40 (1 1/4)",
    },
    "C2": {
        "tipo_caixa": "III",
        "disjuntor_a": 80,
        "medida_eletroduto": "40 (1 1/4)",
    },
    "C3": {
        "tipo_caixa": "III",
        "disjuntor_a": 100,
        "medida_eletroduto": "40 (1 1/4)",
    },
    "C4": {
        "tipo_caixa": "III",
        "disjuntor_a": 125,
        "medida_eletroduto": "50 (1 1/2)",
    },
    "C5": {
        "tipo_caixa": "H",
        "disjuntor_a": 150,
        "medida_eletroduto": "50 (1 1/2)",
    },
    "C6": {
        "tipo_caixa": "H",
        "disjuntor_a": 200,
        "medida_eletroduto": "60 (2)",
    },
}


# -------------------------------
# FORMATACAO
# -------------------------------
def formatar_numero_br(valor: float, casas: int = 1) -> str:
    texto = f"{valor:,.{casas}f}"
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")


def _render_notranslate_guard() -> None:
    st.markdown(
        """
        <meta name="google" content="notranslate">
        <style>
        .notranslate, .notranslate * { translate: no; }
        </style>
        <script>
        (function() {
          const root = document.documentElement;
          const body = document.body;
          if (root) {
            root.setAttribute('translate', 'no');
            root.classList.add('notranslate');
          }
          if (body) {
            body.setAttribute('translate', 'no');
            body.classList.add('notranslate');
          }
        })();
        </script>
        """,
        unsafe_allow_html=True,
    )


def _formatar_df_dimensionamento_exibicao(df: pd.DataFrame | None) -> pd.DataFrame | None:
    if df is None or df.empty:
        return df
    exibicao = df.copy()
    return exibicao.rename(columns={
        'N? circuito': 'N? circuito',
        'Potencia (W/VA)': 'Pot?ncia (W/VA)',
        '1? Criterio': '1? Crit?rio',
        '2? Criterio': '2? Crit?rio',
        '3? Criterio': '3? Crit?rio',
        'Somatoria P x L': 'Somat?ria P x L',
    })


def _formatar_df_pontos_exibicao(df: pd.DataFrame | None) -> pd.DataFrame | None:
    if df is None or df.empty:
        return df
    exibicao = df.copy()
    tipo_labels = {'iluminacao': 'Ilumina??o', 'tug': 'TUG', 'tue': 'TUE'}
    if 'tipo' in exibicao.columns:
        exibicao['tipo'] = exibicao['tipo'].astype(str).map(lambda valor: tipo_labels.get(valor, valor))
    return exibicao.rename(columns={
        'id': 'ID',
        'tipo': 'Tipo',
        'comodo': 'C?modo',
        'item': 'Item',
        'potencia_w': 'Pot?ncia (W)',
        'x': 'X',
        'y': 'Y',
        'circuito': 'Circuito',
        'origem_x': 'Origem X',
        'origem_y': 'Origem Y',
        'distancia_m': 'Dist?ncia (m)',
    })


def _render_df_html_table(df: pd.DataFrame | None) -> None:
    if df is None or df.empty:
        st.info('Sem dados para exibir.')
        return
    html_table = df.to_html(index=False, escape=True, classes='codex-table notranslate', border=0)
    html = f"""<!doctype html>
<html translate="no" class="notranslate">
<head>
<meta charset="utf-8">
<meta name="google" content="notranslate">
<style>
html, body {{
    margin: 0;
    padding: 0;
    background: transparent;
    color: #f9fafb;
    font-family: sans-serif;
}}
.codex-table-wrap {{
    overflow-x: auto;
    border: 1px solid #374151;
    border-radius: 12px;
    background: #111827;
}}
table.codex-table {{
    width: 100%;
    border-collapse: collapse;
    color: #f9fafb;
    font-size: 0.95rem;
}}
table.codex-table thead th {{
    background: #1f2937;
    color: #e5e7eb;
    padding: 10px 12px;
    border-bottom: 1px solid #374151;
    text-align: left;
    white-space: nowrap;
}}
table.codex-table tbody td {{
    padding: 10px 12px;
    border-top: 1px solid #1f2937;
    white-space: nowrap;
}}
table.codex-table tbody tr:nth-child(even) td {{
    background: rgba(255,255,255,0.02);
}}
</style>
</head>
<body translate="no" class="notranslate">
<div class="codex-table-wrap notranslate" translate="no">{html_table}</div>
</body>
</html>"""
    height = min(140 + len(df) * 38, 620)
    components.html(html, height=height, scrolling=True)


# -------------------------------
# NORMALIZACAO
# -------------------------------
def normalizar_ambiente(ambiente: str) -> str:
    amb = ambiente.strip().lower()
    trocas = {
        "a": ["a", "á", "à", "ã", "â"],
        "e": ["e", "é", "ê"],
        "i": ["i", "í"],
        "o": ["o", "ó", "ô", "õ"],
        "u": ["u", "ú"],
        "c": ["c", "ç"],
    }
    for destino, origens in trocas.items():
        for origem in origens[1:]:
            amb = amb.replace(origem, destino)

    aliases = {
        "suite": "suite",
        "suíte": "suite",
        "dormitorio": "quarto",
        "dormitório": "quarto",
        "sala de estar": "sala",
        "sala de jantar": "sala",
        "area de servico": "area_servico",
        "área de servico": "area_servico",
        "area de serviço": "area_servico",
        "área de serviço": "area_servico",
        "copa-cozinha": "copa_cozinha",
        "copa cozinha": "copa_cozinha",
        "bwc": "banheiro",
        "wc": "banheiro",
        "hall de escadaria": "hall",
        "hall escadaria": "hall",
        "casa de maquinas": "casa_maquinas",
        "casa de máquinas": "casa_maquinas",
        "sala de bombas": "sala_bombas",
    }
    return aliases.get(amb, amb)


# -------------------------------
# NBR 5410 - ILUMINACAO
# -------------------------------
def calcular_iluminacao(area: float) -> Dict[str, int]:
    pontos_minimos = 1

    if area <= 6:
        potencia_va = 100
    else:
        acrescimos = math.floor((area - 6) / 4)
        potencia_va = 100 + acrescimos * 60

    pontos_calculados = max(1, math.ceil(potencia_va / 100))
    return {
        "pontos_minimos": pontos_minimos,
        "pontos_calculados": pontos_calculados,
        "potencia_va": potencia_va,
    }


# -------------------------------
# NBR 5410 - TUG
# -------------------------------
def calcular_tug(
    area: float,
    perimetro: float,
    ambiente: str,
    bancadas_validas: int = 0,
) -> Dict[str, object]:
    amb = normalizar_ambiente(ambiente)

    if amb == "banheiro":
        pontos = 1
        potencias = [600]

    elif amb in {"cozinha", "copa", "copa_cozinha", "area_servico", "lavanderia"}:
        pontos_perimetro = math.ceil(perimetro / 3.5)
        pontos = max(pontos_perimetro, bancadas_validas)
        potencias = [600] * min(pontos, 3) + [100] * max(0, pontos - 3)

    elif amb in {"garagem", "varanda", "sotao", "subsolo", "hall", "casa_maquinas", "sala_bombas", "barrilete"}:
        pontos = 1
        potencias = [100]

    else:
        if area <= 6:
            pontos = 1
        else:
            pontos = math.ceil(perimetro / 5)
        potencias = [100] * pontos

    return {
        "pontos": pontos,
        "potencias_va": potencias,
        "potencia_total_va": sum(potencias),
    }


# -------------------------------
# NBR 5410 - TUE
# -------------------------------
def calcular_tue(equipamentos: List[Dict[str, object]]) -> Dict[str, object]:
    total_w = 0.0
    descricoes = []

    for eq in equipamentos:
        nome = str(eq.get("nome", "")).strip() or "Equipamento"
        potencia = float(eq.get("potencia_w", 0) or 0)
        total_w += potencia
        descricoes.append(f"{nome} ({potencia:.0f} W)")

    return {
        "descricao": " / ".join(descricoes) if descricoes else "-",
        "potencia_total_w": total_w,
    }


# -------------------------------
# TUG - FORMATACAO DO PONTO
# -------------------------------
def formatar_sponto_tug(potencias: List[int]) -> str:
    if not potencias:
        return "-"

    contagem = Counter(potencias)
    partes = []
    for va in sorted(contagem.keys(), reverse=True):
        partes.append(f"{contagem[va]} de {va}")
    return " e ".join(partes)


# -------------------------------
# CPFL / GED-13 - DEMANDA (versao simplificada didatica)
# -------------------------------
def fd_iluminacao_tug(carga_kw: float) -> float:
    if carga_kw <= 1:
        return 0.86
    if carga_kw <= 2:
        return 0.75
    if carga_kw <= 3:
        return 0.66
    if carga_kw <= 4:
        return 0.59
    if carga_kw <= 5:
        return 0.52
    if carga_kw <= 6:
        return 0.45
    if carga_kw <= 7:
        return 0.40
    if carga_kw <= 8:
        return 0.35
    if carga_kw <= 9:
        return 0.31
    if carga_kw <= 10:
        return 0.27
    return 0.24


FD_TABELA_B = {
    1: 1.00,
    2: 1.00,
    3: 0.84,
    4: 0.76,
    5: 0.70,
    6: 0.65,
    7: 0.60,
    8: 0.57,
    9: 0.54,
    10: 0.52,
    11: 0.49,
    12: 0.48,
    13: 0.46,
    14: 0.45,
    15: 0.44,
    16: 0.43,
    17: 0.42,
    18: 0.41,
    19: 0.40,
    20: 0.40,
    21: 0.39,
    22: 0.39,
    23: 0.39,
    24: 0.38,
    25: 0.38,
}


def fd_categoria_b(qtd: int) -> float:
    return FD_TABELA_B.get(qtd, 0.38)



def fd_categoria_c(qtd: int) -> float:
    if qtd <= 1:
        return 1.00
    if qtd == 2:
        return 0.72
    return 0.62



def fd_categoria_d(qtd: int) -> float:
    if qtd <= 1:
        return 1.00
    if 2 <= qtd <= 4:
        return 0.70
    if 5 <= qtd <= 6:
        return 0.60
    return 0.50



def fd_categoria_e(qtd: int) -> float:
    tabela = {
        1: 1.00,
        2: 0.60,
        3: 0.48,
        4: 0.40,
        5: 0.37,
        6: 0.35,
        7: 0.33,
        8: 0.32,
        9: 0.31,
    }
    if qtd in tabela:
        return tabela[qtd]
    if 10 <= qtd <= 11:
        return 0.30
    if 12 <= qtd <= 15:
        return 0.28
    return 0.26



def fd_categoria_f_residencial(qtd: int) -> float:
    return 1.00



def demanda_maiores_primeiro(potencias_w: List[float], fatores_maiores: List[float], fator_demais: float) -> float:
    if not potencias_w:
        return 0.0

    potencias = sorted((float(p) for p in potencias_w), reverse=True)
    demanda = 0.0

    for i, potencia in enumerate(potencias):
        if i < len(fatores_maiores):
            demanda += potencia * fatores_maiores[i]
        else:
            demanda += potencia * fator_demais

    return demanda



def calcular_demanda_cpfl_simplificada(
    carga_iluminacao_va: float,
    carga_tug_va: float,
    equipamentos_tue: List[Dict[str, object]],
) -> Tuple[pd.DataFrame, float]:
    linhas = []

    carga_a_kw = (carga_iluminacao_va + carga_tug_va) / 1000
    fd_a = fd_iluminacao_tug(carga_a_kw)
    demanda_a_w = (carga_iluminacao_va + carga_tug_va) * fd_a
    linhas.append(
        {
            "Categoria": "a",
            "Descricao": "Iluminacao + TUG",
            "Carga instalada": f"{carga_iluminacao_va + carga_tug_va:.0f} VA",
            "FD": fd_a,
            "Demanda": f"{demanda_a_w:.0f} W",
            "Equipamentos": "-",
        }
    )

    grupos = defaultdict(list)
    for eq in equipamentos_tue:
        categoria = str(eq.get("categoria_demanda", "x"))
        grupos[categoria].append(eq)

    for categoria, itens in grupos.items():
        potencias = [float(item.get("potencia_w", 0) or 0) for item in itens]
        nomes = " / ".join(str(item.get("nome", "Equipamento")) for item in itens)
        qtd = len(itens)

        if categoria == "b":
            fd = fd_categoria_b(qtd)
            demanda_w = sum(potencias) * fd
            desc = "Chuveiro / torneira / aquecedor de passagem / ferro"
        elif categoria == "c":
            fd = fd_categoria_c(qtd)
            demanda_w = sum(potencias) * fd
            desc = "Boiler / aquecedor central"
        elif categoria == "d":
            fd = fd_categoria_d(qtd)
            demanda_w = sum(potencias) * fd
            desc = "Secadora / forno eletrico / lava-loucas / micro-ondas"
        elif categoria == "e":
            fd = fd_categoria_e(qtd)
            demanda_w = sum(potencias) * fd
            desc = "Fogao eletrico"
        elif categoria == "f":
            fd = fd_categoria_f_residencial(qtd)
            demanda_w = sum(potencias) * fd
            desc = "Ar-condicionado tipo janela (uso residencial)"
        elif categoria == "g":
            fd = "maiores"
            demanda_w = demanda_maiores_primeiro(potencias, [1.00, 0.90, 0.80, 0.80, 0.80], 0.70)
            desc = "Motores / solda a motor"
        elif categoria == "h":
            fd = 1.00
            demanda_w = sum(potencias)
            desc = "Equipamentos especiais (simplificado)"
        elif categoria == "i":
            fd = "maiores"
            demanda_w = demanda_maiores_primeiro(potencias, [1.00, 0.90, 0.80, 0.80, 0.80], 0.70)
            desc = "Hidromassagem"
        else:
            fd = 1.00
            demanda_w = sum(potencias)
            desc = "Sem categoria GED-13"

        linhas.append(
            {
                "Categoria": categoria,
                "Descricao": desc,
                "Carga instalada": f"{sum(potencias):.0f} W",
                "FD": fd,
                "Demanda": f"{demanda_w:.0f} W",
                "Equipamentos": nomes,
            }
        )

    df = pd.DataFrame(linhas)
    total_demanda_w = float(demanda_a_w)
    for _, row in df.iloc[1:].iterrows():
        total_demanda_w += float(str(row["Demanda"]).replace(" W", ""))

    return df, total_demanda_w


# -------------------------------
# PADRAO DE ENTRADA - TABELA 1A GED-13
# -------------------------------
def determinar_categoria_padrao_entrada(carga_instalada_w: float, demanda_total_w: float) -> Tuple[str, str]:
    demanda_total_kva = demanda_total_w / 1000

    if demanda_total_kva <= 6:
        return "A1", "demanda"
    if demanda_total_kva <= 12:
        return "A2", "demanda"
    if demanda_total_kva <= 18:
        return "B1", "demanda"
    if demanda_total_kva <= 25:
        return "B2", "demanda"
    if demanda_total_kva <= 30:
        return "C2", "demanda"
    if demanda_total_kva <= 38:
        return "C3", "demanda"
    if demanda_total_kva <= 47:
        return "C4", "demanda"
    if demanda_total_kva <= 57:
        return "C5", "demanda"
    if demanda_total_kva <= 76:
        return "C6", "demanda"

    return "CONSULTAR GED-13", "consultar GED-13"



def resolver_fase_padrao(categoria: str, fase_escolhida: str) -> str:
    if fase_escolhida != "Automatico":
        return fase_escolhida

    if categoria in {"A1", "A2", "B1", "B2"}:
        return "Monofasico"
    if categoria in {"C1", "C2", "C3", "C4", "C5", "C6"}:
        return "Trifasico"
    return "Consultar GED-13"



def calcular_padrao_entrada(
    carga_instalada_w: float,
    demanda_total_w: float,
    fase_escolhida: str,
) -> Dict[str, str]:
    categoria, criterio = determinar_categoria_padrao_entrada(carga_instalada_w, demanda_total_w)
    fase = resolver_fase_padrao(categoria, fase_escolhida)
    caracteristicas = PADRAO_ENTRADA_TABELA_1A.get(categoria, {})

    return {
        "Fase": fase,
        "Categoria": categoria,
        "Demanda Considerada": f"{formatar_numero_br(demanda_total_w, 1)} W",
        "Carga Instalada": f"{formatar_numero_br(carga_instalada_w, 1)} W",
        "Tipo de Caixa": caracteristicas.get("tipo_caixa", "Consultar GED-13"),
        "Disjuntor": str(caracteristicas.get("disjuntor_a", "Consultar GED-13")),
        "Medida do Eletroduto": caracteristicas.get("medida_eletroduto", "Consultar GED-13"),
        "Criterio da Categoria": criterio,
        "Demanda Total Simplificada": f"{formatar_numero_br(demanda_total_w, 1)} W",
    }



DIMENSIONAMENTO_CORES = {
    'quadro': '#ef4444',
    'iluminacao': '#f59e0b',
    'tug': '#22c55e',
    'tue': '#38bdf8',
}

SECOES_PADRAO = [0.5, 0.75, 1.5, 2.5, 4.0, 6.0, 10.0, 16.0, 25.0, 35.0, 50.0]
AMPACIDADE_CABOS = {
    0.5: 6.0,
    0.75: 9.0,
    1.5: 15.5,
    2.5: 21.0,
    4.0: 28.0,
    6.0: 36.0,
    10.0: 50.0,
    16.0: 68.0,
    25.0: 89.0,
    35.0: 111.0,
    50.0: 134.0,
}
DISJUNTORES_PADRAO = [6, 10, 16, 20, 25, 32, 40, 50, 63]
QUEDA_TENSAO_FATOR = 7.13e-5


def _point_in_polygon(point: tuple[float, float], polygon: list[tuple[float, float]]) -> bool:
    x, y = point
    inside = False
    total = len(polygon)
    for i in range(total):
        x1, y1 = polygon[i]
        x2, y2 = polygon[(i + 1) % total]
        intersects = ((y1 > y) != (y2 > y)) and (x < ((x2 - x1) * (y - y1) / ((y2 - y1) or 1e-9) + x1))
        if intersects:
            inside = not inside
    return inside



def _room_center(comodo: Dict[str, object]) -> tuple[float, float]:
    if comodo.get('centroide_x') is not None and comodo.get('centroide_y') is not None:
        return float(comodo['centroide_x']), float(comodo['centroide_y'])
    vertices = comodo.get('vertices') or []
    if vertices:
        return _polygon_centroid(vertices)
    return 0.0, 0.0



def _room_bbox(comodo: Dict[str, object]) -> tuple[float, float, float, float]:
    vertices = comodo.get('vertices') or []
    if not vertices:
        cx, cy = _room_center(comodo)
        return cx - 1.0, cx + 1.0, cy - 1.0, cy + 1.0
    xs = [float(x) for x, _ in vertices]
    ys = [float(y) for _, y in vertices]
    return min(xs), max(xs), min(ys), max(ys)



def _candidate_offsets() -> list[tuple[float, float]]:
    return [
        (0.0, 0.0),
        (-0.18, 0.0), (0.18, 0.0), (0.0, -0.18), (0.0, 0.18),
        (-0.28, 0.0), (0.28, 0.0), (-0.38, 0.0), (0.38, 0.0),
        (0.0, -0.28), (0.0, 0.28),
    ]



def _point_inside_room(point: tuple[float, float], comodo: Dict[str, object]) -> bool:
    vertices = comodo.get('vertices') or []
    if not vertices:
        return True
    return _point_in_polygon(point, vertices)



def _suggest_lighting_points(comodo: Dict[str, object], count: int) -> list[tuple[float, float]]:
    if count <= 0:
        return []
    cx, cy = _room_center(comodo)
    min_x, max_x, min_y, max_y = _room_bbox(comodo)
    width = max(max_x - min_x, 0.8)
    height = max(max_y - min_y, 0.8)

    if count == 1:
        return [(round(cx, 4), round(cy, 4))]

    points: list[tuple[float, float]] = []
    horizontal = width >= height
    span = (width if horizontal else height) * 0.5
    if count == 2:
        offsets = [-0.5, 0.5]
    elif count == 3:
        offsets = [-0.6, 0.0, 0.6]
    else:
        offsets = [(-1.0 + 2.0 * i / (count - 1)) for i in range(count)]

    for factor in offsets:
        if horizontal:
            candidate = (cx + factor * span, cy)
        else:
            candidate = (cx, cy + factor * span)
        if not _point_inside_room(candidate, comodo):
            candidate = (cx + (candidate[0] - cx) * 0.6, cy + (candidate[1] - cy) * 0.6)
        points.append((round(candidate[0], 4), round(candidate[1], 4)))
    return points



def _suggest_wall_points(comodo: Dict[str, object], count: int) -> list[tuple[float, float]]:
    if count <= 0:
        return []
    min_x, max_x, min_y, max_y = _room_bbox(comodo)
    cx, cy = _room_center(comodo)
    width = max(max_x - min_x, 0.8)
    height = max(max_y - min_y, 0.8)
    inset_x = width * 0.08
    inset_y = height * 0.08

    wall_candidates = [
        (min_x + inset_x, cy),
        (max_x - inset_x, cy),
        (cx, max_y - inset_y),
        (cx, min_y + inset_y),
        (min_x + inset_x, max_y - inset_y),
        (max_x - inset_x, max_y - inset_y),
        (min_x + inset_x, min_y + inset_y),
        (max_x - inset_x, min_y + inset_y),
    ]

    points: list[tuple[float, float]] = []
    for candidate in wall_candidates:
        point = candidate
        if not _point_inside_room(point, comodo):
            point = ((candidate[0] + cx) / 2.0, (candidate[1] + cy) / 2.0)
        if not _point_inside_room(point, comodo):
            point = (cx, cy)
        point = (round(point[0], 4), round(point[1], 4))
        if point not in points:
            points.append(point)
        if len(points) >= count:
            return points

    while len(points) < count:
        points.append((round(cx, 4), round(cy, 4)))
    return points



def _suggest_points_for_room(comodo: Dict[str, object], count: int, band: str = 'center') -> list[tuple[float, float]]:
    if band == 'top':
        return _suggest_lighting_points(comodo, count)
    if band == 'center':
        return _suggest_wall_points(comodo, count)
    return _suggest_wall_points(comodo, count)



def _room_anchor_point(comodo: Dict[str, object], modo: str) -> tuple[float, float]:
    cx, cy = _room_center(comodo)
    min_x, max_x, min_y, max_y = _room_bbox(comodo)
    width = max(max_x - min_x, 0.8)
    height = max(max_y - min_y, 0.8)
    inset_x = width * 0.1
    inset_y = height * 0.1
    left_x = min_x + inset_x
    right_x = max_x - inset_x
    bottom_y = min_y + inset_y
    top_y = max_y - inset_y
    anchors = {
        'Centro': (cx, cy),
        'Esquerda': (left_x, cy),
        'Canto esquerda superior': (left_x, top_y),
        'Canto esquerda inferior': (left_x, bottom_y),
        'Cima': (cx, top_y),
        'Canto cima direita': (right_x, top_y),
        'Canto cima esquerda': (left_x, top_y),
        'Baixo': (cx, bottom_y),
        'Canto baixo direita': (right_x, bottom_y),
        'Canto baixo esquerda': (left_x, bottom_y),
        'Direita': (right_x, cy),
        'Canto direita superior': (right_x, top_y),
        'Canto direita inferior': (right_x, bottom_y),
    }
    point = anchors.get(modo, (cx, cy))
    if not _point_inside_room(point, comodo):
        point = (cx, cy)
    return round(point[0], 4), round(point[1], 4)

def _find_comodo_by_name(comodos: List[Dict[str, object]], nome: str) -> Dict[str, object] | None:
    return next((comodo for comodo in comodos if str(comodo.get('nome', '')) == nome), None)



def _label_tipo_ponto(tipo: str) -> str:
    labels = {
        'iluminacao': 'Iluminacao',
        'tug': 'Tomada',
        'tue': 'Equipamento TUE',
    }
    return labels.get(tipo, tipo.title())



def _position_options_for_point(tipo: str) -> List[str]:
    return [
        'Centro',
        'Esquerda',
        'Canto esquerda superior',
        'Canto esquerda inferior',
        'Cima',
        'Canto cima direita',
        'Canto cima esquerda',
        'Baixo',
        'Canto baixo direita',
        'Canto baixo esquerda',
        'Direita',
        'Canto direita superior',
        'Canto direita inferior',
    ]

def _suggest_tug_circuit_label(comodo: Dict[str, object]) -> str:
    tipo = normalizar_ambiente(str(comodo.get('tipo', 'outro')))
    nome = str(comodo.get('nome', 'Comodo')).upper()
    if tipo in {'suite', 'quarto', 'corredor', 'circulacao', 'hall', 'closet'}:
        return 'TUG (SUITE | QUARTO | CORREDOR)'
    if tipo == 'sala':
        return 'TUG (SALA / ESTAR / JANTAR)'
    if tipo in {'cozinha', 'copa', 'copa_cozinha'}:
        return 'TUG - COZINHA'
    if tipo in {'area_servico', 'lavanderia'}:
        return 'TUG - AREA DE SERVICO'
    if tipo == 'banheiro':
        return 'TUG - BANHEIRO'
    if tipo == 'lavabo':
        return 'TUG - LAVABO'
    if tipo == 'varanda':
        return 'TUG - VARANDA'
    return f'TUG - {nome}'



def _suggest_tue_circuit_label(comodo: Dict[str, object], equipamento: Dict[str, object]) -> str:
    tipo = normalizar_ambiente(str(comodo.get('tipo', 'outro')))
    nome = str(comodo.get('nome', 'Comodo')).upper()
    equipamento_nome = str(equipamento.get('nome', 'Equipamento')).strip().upper()
    if tipo in {'cozinha', 'copa', 'copa_cozinha'}:
        return 'TUE - COZINHA'
    if tipo in {'area_servico', 'lavanderia'}:
        return 'TUE - AREA DE SERVICO'
    if tipo == 'banheiro':
        return 'TUE - BANHEIRO'
    if tipo == 'suite':
        return 'TUE - SUITE'
    if tipo == 'quarto':
        return 'TUE - QUARTO'
    return f'TUE - {nome} ({equipamento_nome})'



def _next_standard_section(value: float) -> float:
    for section in SECOES_PADRAO:
        if section >= value - 1e-9:
            return section
    return SECOES_PADRAO[-1]



def _next_standard_breaker(current: float) -> int:
    current = max(current, 0.1)
    for breaker in DISJUNTORES_PADRAO:
        if breaker >= current - 1e-9:
            return breaker
    return DISJUNTORES_PADRAO[-1]



def _section_by_ampacity(corrected_current: float) -> float:
    for section, ampacity in AMPACIDADE_CABOS.items():
        if ampacity >= corrected_current - 1e-9:
            return section
    return max(AMPACIDADE_CABOS)



def _section_by_short_circuit(breaker: int) -> float:
    if breaker >= 63:
        return 16.0
    if breaker >= 50:
        return 10.0
    if breaker >= 32:
        return 6.0
    return 4.0



def _distance_orthogonal(a: tuple[float, float], b: tuple[float, float]) -> float:
    return abs(a[0] - b[0]) + abs(a[1] - b[1])



def _expected_dimensionamento_point_count(comodos: List[Dict[str, object]]) -> int:
    total = 0
    for comodo in comodos:
        total += max(1, int(comodo.get('iluminacao_pontos', 1) or 1))
        total += len(comodo.get('tug_potencias_va', []) or [])
        total += len(comodo.get('tues', []) or [])
    return total


def _sanitize_dimensionamento_points(
    pontos: List[Dict[str, object]],
    comodos: List[Dict[str, object]],
    endpoints: list[tuple[float, float]],
) -> tuple[dict[str, float] | None, List[Dict[str, object]]]:
    allowed_types = {'iluminacao', 'tug', 'tue'}
    room_names = {str(comodo.get('nome', 'Comodo')) for comodo in comodos}
    expected_count = _expected_dimensionamento_point_count(comodos)
    has_tue = any(comodo.get('tues') for comodo in comodos)

    invalid = len(pontos) != expected_count
    if not invalid:
        for ponto in pontos:
            tipo = str(ponto.get('tipo', ''))
            item = str(ponto.get('item', ''))
            comodo_nome = str(ponto.get('comodo', ''))
            if tipo not in allowed_types:
                invalid = True
                break
            if comodo_nome not in room_names:
                invalid = True
                break
            if tipo == 'iluminacao' and item != 'Ponto de luz':
                invalid = True
                break
            if tipo == 'tug' and not (item.startswith('TUG ') or item.startswith('Tomada ')):
                invalid = True
                break
            if tipo == 'tue' and not has_tue:
                invalid = True
                break

    if invalid:
        return _build_dimensionamento_points(comodos, endpoints)
    return None, pontos


def _build_dimensionamento_points(comodos: List[Dict[str, object]], endpoints: list[tuple[float, float]]) -> tuple[dict[str, float], list[dict[str, object]]]:
    view = _default_view_state(endpoints)
    quadro = {
        'x': round(view['center_x'], 4),
        'y': round(view['center_y'], 4),
    }
    points: list[dict[str, object]] = []
    point_idx = 1

    for comodo in comodos:
        iluminacao_pontos = max(1, int(comodo.get('iluminacao_pontos', 1)))
        ilum_positions = _suggest_points_for_room(comodo, iluminacao_pontos, band='top')
        ilum_power = float(comodo.get('iluminacao_stotal_va', 0) or 0)
        power_per_point = ilum_power / max(len(ilum_positions), 1)
        for pos in ilum_positions:
            points.append({
                'id': f'P{point_idx:03d}',
                'tipo': 'iluminacao',
                'comodo': str(comodo.get('nome', 'Comodo')),
                'item': 'Ponto de luz',
                'potencia_w': round(power_per_point, 2),
                'x': pos[0],
                'y': pos[1],
                'circuito': 'ILUMINACAO',
                'origem_x': pos[0],
                'origem_y': pos[1],
            })
            point_idx += 1

        tug_powers = [float(p) for p in comodo.get('tug_potencias_va', [])]
        tug_positions = _suggest_points_for_room(comodo, len(tug_powers), band='center')
        tug_circuit = _suggest_tug_circuit_label(comodo)
        for idx, potencia in enumerate(tug_powers):
            pos = tug_positions[idx] if idx < len(tug_positions) else _room_center(comodo)
            points.append({
                'id': f'P{point_idx:03d}',
                'tipo': 'tug',
                'comodo': str(comodo.get('nome', 'Comodo')),
                'item': f'Tomada {idx + 1}',
                'potencia_w': round(potencia, 2),
                'x': pos[0],
                'y': pos[1],
                'circuito': tug_circuit,
                'origem_x': pos[0],
                'origem_y': pos[1],
            })
            point_idx += 1

        tue_positions = _suggest_points_for_room(comodo, len(comodo.get('tues', [])), band='bottom')
        for idx, equipamento in enumerate(comodo.get('tues', [])):
            pos = tue_positions[idx] if idx < len(tue_positions) else _room_center(comodo)
            points.append({
                'id': f'P{point_idx:03d}',
                'tipo': 'tue',
                'comodo': str(comodo.get('nome', 'Comodo')),
                'item': str(equipamento.get('nome', f'TUE {idx + 1}')),
                'potencia_w': round(float(equipamento.get('potencia_w', 0) or 0), 2),
                'x': pos[0],
                'y': pos[1],
                'circuito': _suggest_tue_circuit_label(comodo, equipamento),
                'origem_x': pos[0],
                'origem_y': pos[1],
            })
            point_idx += 1

    return quadro, points



def _build_dimensionamento_figure(
    segments: list[tuple[tuple[float, float], tuple[float, float]]],
    endpoints: list[tuple[float, float]],
    comodos: List[Dict[str, object]],
    quadro: dict[str, float],
    pontos: List[Dict[str, object]],
    show_paths: bool,
    view_state: dict[str, float] | None = None,
) -> go.Figure:
    effective_view = view_state or _default_view_state(endpoints)
    background_image, transform = _build_canvas_background(segments, endpoints, effective_view)
    fig = go.Figure()

    img_buffer = io.BytesIO()
    background_image.save(img_buffer, format='PNG')
    encoded = base64.b64encode(img_buffer.getvalue()).decode('ascii')
    image_source = f'data:image/png;base64,{encoded}'

    fig.update_layout(
        images=[dict(
            source=image_source,
            xref='x',
            yref='y',
            x=0,
            y=0,
            sizex=transform['canvas_width'],
            sizey=transform['canvas_height'],
            sizing='stretch',
            layer='below',
        )],
        height=760,
        margin=dict(l=10, r=10, t=10, b=10),
        paper_bgcolor='#111827',
        plot_bgcolor='#1F2937',
        legend=dict(orientation='h', yanchor='bottom', y=1.01, xanchor='left', x=0),
    )

    for comodo in comodos:
        vertices = comodo.get('vertices') or []
        if len(vertices) >= 3:
            canvas_vertices = [_world_to_canvas_point((float(x), float(y)), transform) for x, y in vertices]
            xs = [p[0] for p in canvas_vertices] + [canvas_vertices[0][0]]
            ys = [p[1] for p in canvas_vertices] + [canvas_vertices[0][1]]
            fig.add_trace(
                go.Scatter(
                    x=xs,
                    y=ys,
                    fill='toself',
                    mode='lines',
                    line=dict(color='rgba(16, 185, 129, 0.95)', width=2),
                    fillcolor='rgba(16, 185, 129, 0.16)',
                    name=str(comodo.get('nome', 'Comodo')),
                    hovertemplate=f"{comodo.get('nome', 'Comodo')}<br>Area: {comodo.get('area', 0):.2f} m2<extra></extra>",
                    showlegend=False,
                )
            )
        cx, cy = _world_to_canvas_point(_room_center(comodo), transform)
        fig.add_trace(
            go.Scatter(
                x=[cx],
                y=[cy],
                mode='text',
                text=[str(comodo.get('nome', 'Comodo'))],
                textfont=dict(color='#93C5FD', size=11),
                hoverinfo='skip',
                showlegend=False,
            )
        )

    quadro_canvas = _world_to_canvas_point((float(quadro['x']), float(quadro['y'])), transform)
    if show_paths:
        for ponto in pontos:
            ponto_canvas = _world_to_canvas_point((float(ponto['x']), float(ponto['y'])), transform)
            fig.add_trace(
                go.Scatter(
                    x=[quadro_canvas[0], ponto_canvas[0], ponto_canvas[0]],
                    y=[quadro_canvas[1], quadro_canvas[1], ponto_canvas[1]],
                    mode='lines',
                    line=dict(color='rgba(251, 191, 36, 0.55)', width=1.5, dash='dot'),
                    hoverinfo='skip',
                    showlegend=False,
                )
            )

    fig.add_trace(
        go.Scatter(
            x=[quadro_canvas[0]],
            y=[quadro_canvas[1]],
            mode='markers+text',
            marker=dict(size=14, color=DIMENSIONAMENTO_CORES['quadro'], symbol='diamond'),
            text=['QD'],
            textposition='top center',
            name='Quadro',
        )
    )

    for tipo, legenda in [('iluminacao', 'Iluminacao'), ('tug', 'TUG'), ('tue', 'TUE')]:
        subset = [p for p in pontos if p['tipo'] == tipo]
        if not subset:
            continue
        coords = [_world_to_canvas_point((float(p['x']), float(p['y'])), transform) for p in subset]
        fig.add_trace(
            go.Scatter(
                x=[p[0] for p in coords],
                y=[p[1] for p in coords],
                mode='markers+text',
                marker=dict(size=11, color=DIMENSIONAMENTO_CORES[tipo]),
                text=[p['id'] for p in subset],
                textposition='top center',
                name=legenda,
                customdata=[[p['comodo'], p['item'], p['circuito']] for p in subset],
                hovertemplate='Comodo: %{customdata[0]}<br>Item: %{customdata[1]}<br>Circuito: %{customdata[2]}<extra></extra>',
            )
        )

    fig.update_xaxes(range=[0, transform['canvas_width']], showgrid=False, visible=False)
    fig.update_yaxes(range=[transform['canvas_height'], 0], showgrid=False, visible=False, scaleanchor='x', scaleratio=1)
    return fig



def _compute_dimensionamento_tables(
    pontos: List[Dict[str, object]],
    quadro: dict[str, float],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    pontos_com_distancia: list[dict[str, object]] = []
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)

    for ponto in pontos:
        distancia = _distance_orthogonal((quadro['x'], quadro['y']), (float(ponto['x']), float(ponto['y'])))
        registro = dict(ponto)
        registro['distancia_m'] = round(distancia, 2)
        pontos_com_distancia.append(registro)
        grouped[str(ponto['circuito'])].append(registro)

    rows = []
    ordered_circuits = list(grouped.keys())
    for idx, circuito in enumerate(ordered_circuits, start=1):
        itens = grouped[circuito]
        total_potencia = sum(float(item['potencia_w']) for item in itens)
        pl_sum = sum(float(item['potencia_w']) * float(item['distancia_m']) for item in itens)
        distancias = [float(item['distancia_m']) for item in itens]
        distancia_expr = ' + '.join(f'{valor:.2f}' for valor in distancias)
        criterio_1 = 1.5 if all(item['tipo'] == 'iluminacao' for item in itens) else 2.5
        tensao = 127.0
        ib = total_potencia / tensao if tensao else 0.0
        fator_correcao = 0.94 * (0.8 if all(item['tipo'] in {'iluminacao', 'tug'} for item in itens) else 1.0)
        ib_corrigida = ib / fator_correcao if fator_correcao else ib
        criterio_2 = _section_by_ampacity(ib_corrigida)
        criterio_3 = pl_sum * QUEDA_TENSAO_FATOR
        disjuntor = _next_standard_breaker(ib)
        curto_circuito = _section_by_short_circuit(disjuntor)
        cabo_escolhido = _next_standard_section(max(criterio_1, criterio_2, criterio_3, curto_circuito))

        rows.append({
            'N? circuito': idx,
            'Circuitos': circuito,
            'Potencia (W/VA)': round(total_potencia, 1),
            'Distancia': distancia_expr,
            '1? Criterio': criterio_1,
            '2? Criterio': criterio_2,
            '3? Criterio': round(criterio_3, 3),
            'Disjuntores': disjuntor,
            'Curto Circuito': curto_circuito,
            'Cabo Escolhido': cabo_escolhido,
            'IB': round(ib, 3),
            "IB'": round(ib_corrigida, 3),
            'Somatoria P x L': round(pl_sum, 2),
        })

    df_dimensionamento = pd.DataFrame(rows)
    df_pontos = pd.DataFrame(pontos_com_distancia)
    return df_dimensionamento, df_pontos



def montar_projeto_calculado(
    nome_projeto: str,
    responsavel: str,
    fase_padrao: str,
    comodos: List[Dict[str, object]],
    equipamentos_gerais_demanda: List[Dict[str, object]],
    segments: list[tuple[tuple[float, float], tuple[float, float]]] | None,
    endpoints: list[tuple[float, float]] | None,
    comodos_importados: List[Dict[str, object]] | None = None,
) -> Dict[str, object]:
    resultados = []
    comodos_calculados: list[dict[str, object]] = []
    total_ilum_va = 0.0
    total_tug_va = 0.0
    total_tue_w = 0.0

    comodos_importados = comodos_importados or []

    for idx, comodo in enumerate(comodos):
        comodo_base = dict(comodo)
        if idx < len(comodos_importados):
            importado = comodos_importados[idx]
            if not comodo_base.get('vertices'):
                comodo_base['vertices'] = importado.get('vertices', [])
                comodo_base['centroide_x'] = importado.get('centroide_x')
                comodo_base['centroide_y'] = importado.get('centroide_y')
                comodo_base['canvas_vertices'] = importado.get('canvas_vertices', [])
                comodo_base['capture_transform'] = importado.get('capture_transform')
        comodo_base = _normalize_room_geometry(comodo_base, endpoints or [])
        ilum = calcular_iluminacao(comodo_base['area'])
        tug = calcular_tug(
            area=comodo_base['area'],
            perimetro=comodo_base['perimetro'],
            ambiente=comodo_base['tipo'],
            bancadas_validas=int(comodo_base['bancadas_validas']),
        )
        tue = calcular_tue(comodo_base['tues'])

        total_ilum_va += ilum['potencia_va']
        total_tug_va += tug['potencia_total_va']
        total_tue_w += tue['potencia_total_w']

        resultados.append(
            {
                'Comodo': comodo['nome'],
                'Tipo': comodo['tipo'],
                'Area (m2)': round(comodo['area'], 2),
                'Perimetro (m)': round(comodo['perimetro'], 2),
                'Iluminacao - Pontos min.': ilum['pontos_calculados'],
                'Iluminacao - Stotal (VA)': ilum['potencia_va'],
                'TUG - Pontos': tug['pontos'],
                'TUG - Sponto (VA)': formatar_sponto_tug(tug['potencias_va']),
                'TUG - Stotal (VA)': tug['potencia_total_va'],
                'TUE - Equipamentos': tue['descricao'],
                'TUE - Ptotal (W)': round(tue['potencia_total_w'], 1),
            }
        )

        comodos_calculados.append({
            **comodo_base,
            'iluminacao_pontos': ilum['pontos_calculados'],
            'iluminacao_stotal_va': ilum['potencia_va'],
            'tug_pontos': tug['pontos'],
            'tug_potencias_va': tug['potencias_va'],
            'tug_stotal_va': tug['potencia_total_va'],
            'tue_descricao': tue['descricao'],
            'tue_total_w': tue['potencia_total_w'],
        })

    df_resultados = pd.DataFrame(resultados)
    potencia_instalada_total_w = total_ilum_va + total_tug_va + total_tue_w
    df_demanda, total_demanda_w = calcular_demanda_cpfl_simplificada(
        carga_iluminacao_va=total_ilum_va,
        carga_tug_va=total_tug_va,
        equipamentos_tue=equipamentos_gerais_demanda,
    )
    padrao_entrada = calcular_padrao_entrada(
        carga_instalada_w=potencia_instalada_total_w,
        demanda_total_w=total_demanda_w,
        fase_escolhida=fase_padrao,
    )

    df_padrao = pd.DataFrame(
        {
            'Caracteristica': [
                'Fase',
                'Categoria',
                'Demanda Considerada',
                'Carga Instalada',
                'Tipo de Caixa',
                'Disjuntor',
                'Medida do Eletroduto',
            ],
            'Valor': [
                padrao_entrada['Fase'],
                padrao_entrada['Categoria'],
                padrao_entrada['Demanda Considerada'],
                padrao_entrada['Carga Instalada'],
                padrao_entrada['Tipo de Caixa'],
                padrao_entrada['Disjuntor'],
                padrao_entrada['Medida do Eletroduto'],
            ],
        }
    )

    df_resumo = pd.DataFrame(
        {
            'Indicador': [
                'Projeto',
                'Responsavel',
                'Quantidade de comodos',
                'Iluminacao total (VA)',
                'TUG total (VA)',
                'TUE total (W)',
                'Carga instalada total (W)',
                'Demanda total simplificada (W)',
                'Fase padrao',
                'Categoria padrao',
                'Tipo de caixa',
                'Disjuntor (A)',
                'Medida do eletroduto',
            ],
            'Valor': [
                nome_projeto,
                responsavel or '-',
                len(comodos_calculados),
                total_ilum_va,
                total_tug_va,
                total_tue_w,
                potencia_instalada_total_w,
                total_demanda_w,
                padrao_entrada['Fase'],
                padrao_entrada['Categoria'],
                padrao_entrada['Tipo de Caixa'],
                padrao_entrada['Disjuntor'],
                padrao_entrada['Medida do Eletroduto'],
            ],
        }
    )

    assinatura = _hash_bytes(json.dumps({
        'nome': nome_projeto,
        'responsavel': responsavel,
        'comodos': comodos_calculados,
    }, ensure_ascii=False, sort_keys=True, default=str).encode('utf-8'))

    return {
        'assinatura': assinatura,
        'nome_projeto': nome_projeto,
        'responsavel': responsavel,
        'fase_padrao': fase_padrao,
        'comodos': comodos_calculados,
        'segments': segments or [],
        'endpoints': endpoints or [],
        'view_state': st.session_state.get(_session_key('view_state')) if endpoints else None,
        'df_resultados': df_resultados,
        'df_demanda': df_demanda,
        'df_padrao': df_padrao,
        'df_resumo': df_resumo,
        'padrao_entrada': padrao_entrada,
        'total_ilum_va': total_ilum_va,
        'total_tug_va': total_tug_va,
        'total_tue_w': total_tue_w,
        'potencia_instalada_total_w': potencia_instalada_total_w,
        'total_demanda_w': total_demanda_w,
    }



def renderizar_dimensionamento(projeto: Dict[str, object]) -> tuple[pd.DataFrame | None, pd.DataFrame | None]:
    segments = projeto.get('segments') or []
    endpoints = projeto.get('endpoints') or []
    projeto['comodos'] = [_normalize_room_geometry(comodo, endpoints) for comodo in projeto.get('comodos', [])]
    if not segments or not endpoints:
        st.warning('O dimensionamento visual precisa de um projeto importado por DXF.')
        return None, None

    assinatura = str(projeto['assinatura'])
    quadro_key = f'dimensionamento_quadro_{assinatura}'
    pontos_key = f'dimensionamento_pontos_{assinatura}'
    resultado_key = f'dimensionamento_resultado_{assinatura}'
    finalizado_key = f'dimensionamento_finalizado_{assinatura}'
    quadro_request_key = f'dimensionamento_quadro_request_{assinatura}'
    ponto_request_key = f'dimensionamento_ponto_request_{assinatura}'

    if quadro_key not in st.session_state or pontos_key not in st.session_state:
        quadro, pontos = _build_dimensionamento_points(projeto['comodos'], endpoints)
        st.session_state[quadro_key] = quadro
        st.session_state[pontos_key] = pontos
    else:
        pontos_sessao = list(st.session_state.get(pontos_key, []))
        quadro_rebuild, pontos_sanitizados = _sanitize_dimensionamento_points(pontos_sessao, projeto['comodos'], endpoints)
        if quadro_rebuild is not None:
            st.session_state[quadro_key] = quadro_rebuild
            st.session_state[pontos_key] = pontos_sanitizados
            st.session_state.pop(resultado_key, None)
            st.session_state[finalizado_key] = False

    applied_request = False
    quadro_request = st.session_state.pop(quadro_request_key, None)
    if quadro_request:
        comodo_alvo = _find_comodo_by_name(projeto['comodos'], str(quadro_request['comodo']))
        if comodo_alvo is not None:
            xq, yq = _room_anchor_point(comodo_alvo, str(quadro_request['posicao']))
            st.session_state[quadro_key] = {'x': xq, 'y': yq}
            st.session_state.pop(resultado_key, None)
            st.session_state[finalizado_key] = False
            applied_request = True

    ponto_request = st.session_state.pop(ponto_request_key, None)
    if ponto_request:
        pontos_tmp = list(st.session_state.get(pontos_key, []))
        ponto_id = str(ponto_request['id'])
        comodo_nome = str(ponto_request['comodo'])
        posicao = str(ponto_request['posicao'])
        comodo_ref = _find_comodo_by_name(projeto['comodos'], comodo_nome)
        if comodo_ref is not None:
            ax, ay = _room_anchor_point(comodo_ref, posicao)
            for ponto_tmp in pontos_tmp:
                if str(ponto_tmp.get('id')) == ponto_id:
                    ponto_tmp['x'] = ax
                    ponto_tmp['y'] = ay
                    break
            st.session_state[pontos_key] = pontos_tmp
            st.session_state.pop(resultado_key, None)
            st.session_state[finalizado_key] = False
            applied_request = True

    if applied_request:
        st.rerun()

    st.subheader('8) Dimensionamento')
    st.caption('O quadro e os pontos sugeridos podem ser ajustados. As distancias usam trajeto ortogonal simplificado |dx| + |dy|.')

    if st.session_state.get(finalizado_key):
        st.success('Dimensionamento concluido. Voce pode exportar o XLSX ou voltar para edicao.')
        resultado = st.session_state.get(resultado_key)
        if st.button('Voltar para edicao do dimensionamento', key=f'editar_dim_{assinatura}', use_container_width=True):
            st.session_state[finalizado_key] = False
            st.rerun()
        if not resultado:
            return None, None
        df_dimensionamento, df_pontos = resultado
        st.markdown('**Tabela de dimensionamento**')
        _render_df_html_table(_formatar_df_dimensionamento_exibicao(df_dimensionamento))
        st.markdown('**Pontos e dist?ncias**')
        _render_df_html_table(_formatar_df_pontos_exibicao(df_pontos))
        return df_dimensionamento, df_pontos

    action_cols = st.columns([1, 1, 1])
    with action_cols[0]:
        if st.button('Recarregar pontos sugeridos', key=f'reset_dim_{assinatura}', use_container_width=True):
            quadro, pontos = _build_dimensionamento_points(projeto['comodos'], endpoints)
            st.session_state[quadro_key] = quadro
            st.session_state[pontos_key] = pontos
            st.session_state.pop(resultado_key, None)
            st.rerun()
    with action_cols[1]:
        show_paths = st.toggle('Mostrar percursos', value=True, key=f'show_paths_{assinatura}')
    with action_cols[2]:
        st.info('Quadro e TUE podem ser ajustados por posicao presetada.')

    quadro = dict(st.session_state[quadro_key])
    pontos = list(st.session_state[pontos_key])

    q_col1, q_col2, q_col3 = st.columns([1.2, 1.2, 1])
    with q_col1:
        opcoes_quadro = [str(c['nome']) for c in projeto['comodos']]
        alvo_quadro = st.selectbox('Comodo do quadro', options=opcoes_quadro, key=f'qd_target_{assinatura}')
    with q_col2:
        modo_quadro = st.selectbox(
            'Posicao do quadro',
            options=_position_options_for_point('quadro'),
            key=f'qd_mode_{assinatura}',
        )
    with q_col3:
        if st.button('Aplicar quadro', key=f'qd_apply_{assinatura}', use_container_width=True):
            st.session_state[quadro_request_key] = {'comodo': alvo_quadro, 'posicao': modo_quadro}
            st.session_state.pop(resultado_key, None)
            st.session_state[finalizado_key] = False
            st.rerun()
    st.caption(f"Quadro atual: X={quadro['x']:.2f} | Y={quadro['y']:.2f}")
    st.session_state[quadro_key] = quadro

    figura = _build_dimensionamento_figure(segments, endpoints, projeto['comodos'], quadro, pontos, show_paths, projeto.get('view_state'))
    st.plotly_chart(figura, use_container_width=True, config={'scrollZoom': True, 'displaylogo': False})

    st.markdown('**Editar pontos do dimensionamento**')
    st.caption('Iluminacao e TUG permanecem nas posicoes autom?ticas. Ajuste manualmente apenas TUE, se necessario.')
    tipo_filtro = st.selectbox(
        'Filtro de pontos',
        options=['Todos', 'tue'],
        key=f'filtro_pontos_{assinatura}',
    )
    pontos_editados = []
    aplicar_posicao_em_algum = False
    for idx, ponto in enumerate(pontos):
        ponto_atualizado = dict(ponto)
        if ponto['tipo'] != 'tue':
            pontos_editados.append(ponto_atualizado)
            continue
        if tipo_filtro != 'Todos' and ponto['tipo'] != tipo_filtro:
            pontos_editados.append(ponto_atualizado)
            continue
        with st.expander(f"{ponto['id']} | {_label_tipo_ponto(str(ponto['tipo']))} | {ponto['comodo']} | {ponto['item']}", expanded=False):
            st.caption(f"Posicao atual: X={float(ponto['x']):.2f} | Y={float(ponto['y']):.2f}")
            col_a, col_b = st.columns([1.3, 1])
            with col_a:
                novo_circuito = st.text_input(
                    f"Circuito {ponto['id']}",
                    value=str(ponto['circuito']),
                    key=f"pc_{assinatura}_{idx}",
                )
            with col_b:
                preset_posicao = st.selectbox(
                    f"Posicao {ponto['id']}",
                    options=_position_options_for_point('tue'),
                    key=f"pp_{assinatura}_{idx}",
                )
                aplicar_posicao = st.button(
                    f"Aplicar {ponto['id']}",
                    key=f"ppa_{assinatura}_{idx}",
                    use_container_width=True,
                )
            if aplicar_posicao:
                st.session_state[ponto_request_key] = {
                    'id': ponto['id'],
                    'comodo': ponto['comodo'],
                    'posicao': preset_posicao,
                }
                st.session_state.pop(resultado_key, None)
                st.session_state[finalizado_key] = False
                aplicar_posicao_em_algum = True
            ponto_atualizado['circuito'] = novo_circuito.strip() or str(ponto['circuito'])
            pontos_editados.append(ponto_atualizado)

    pontos_editados.sort(key=lambda p: p['id'])
    st.session_state[pontos_key] = pontos_editados
    if st.button('Concluir dimensionamento', key=f'concluir_dim_{assinatura}', type='primary'):
        st.session_state[resultado_key] = _compute_dimensionamento_tables(st.session_state[pontos_key], quadro)
        st.session_state[finalizado_key] = True
        st.rerun()

    return None, None


# -------------------------------
# EXPORTACAO XLSX
# -------------------------------
def ajustar_largura_colunas(worksheet) -> None:
    for coluna in worksheet.columns:
        maximo = 0
        indice_coluna = coluna[0].column
        letra = get_column_letter(indice_coluna)
        for celula in coluna:
            valor = "" if celula.value is None else str(celula.value)
            if len(valor) > maximo:
                maximo = len(valor)
        worksheet.column_dimensions[letra].width = min(maximo + 3, 60)



def aplicar_estilo_planilha(writer: pd.ExcelWriter) -> None:
    workbook = writer.book
    cabecalho_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    cabecalho_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        if worksheet.max_row >= 1:
            for celula in worksheet[1]:
                celula.fill = cabecalho_fill
                celula.font = cabecalho_font
        worksheet.freeze_panes = "A2"
        ajustar_largura_colunas(worksheet)



def gerar_excel_bytes(
    df_resultados: pd.DataFrame,
    df_demanda: pd.DataFrame,
    df_padrao: pd.DataFrame,
    df_resumo: pd.DataFrame,
    df_dimensionamento: pd.DataFrame | None = None,
    df_pontos_dimensionamento: pd.DataFrame | None = None,
) -> bytes:
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_resultados.to_excel(writer, sheet_name="Tabela de Cargas", index=False)
        df_demanda.to_excel(writer, sheet_name="Demanda CPFL", index=False)
        df_padrao.to_excel(writer, sheet_name="Padrao de Entrada", index=False)
        df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
        if df_dimensionamento is not None and not df_dimensionamento.empty:
            df_dimensionamento.to_excel(writer, sheet_name="Dimensionamento", index=False)
        if df_pontos_dimensionamento is not None and not df_pontos_dimensionamento.empty:
            df_pontos_dimensionamento.to_excel(writer, sheet_name="Pontos Distancias", index=False)
        aplicar_estilo_planilha(writer)

    buffer.seek(0)
    return buffer.getvalue()


ARQUIVO_SECRETS = Path(__file__).with_name("secrets.toml")
ARQUIVO_OAUTH_ESTADO = Path(__file__).with_name("oauth_state.json")


# -------------------------------
# AUTENTICACAO
# -------------------------------
def carregar_config_autenticacao() -> Dict[str, object]:
    config: Dict[str, object] = {}

    try:
        config.update(dict(st.secrets))
    except Exception:
        pass

    if ARQUIVO_SECRETS.exists() and tomllib is not None:
        with ARQUIVO_SECRETS.open("rb") as arquivo:
            config_arquivo = tomllib.load(arquivo)
        for chave, valor in config_arquivo.items():
            config[chave] = valor

    return config


def obter_query_params() -> Dict[str, str]:
    if hasattr(st, "query_params"):
        params = {}
        for chave in st.query_params.keys():
            valor = st.query_params[chave]
            if isinstance(valor, list):
                params[chave] = valor[0] if valor else ""
            else:
                params[chave] = valor
        return params

    params_legacy = st.experimental_get_query_params()
    return {chave: valores[0] for chave, valores in params_legacy.items() if valores}


def limpar_query_params() -> None:
    if hasattr(st, "query_params"):
        for chave in list(st.query_params.keys()):
            del st.query_params[chave]
    else:
        st.experimental_set_query_params()


def salvar_estado_oauth(state: str, provider: str) -> None:
    dados = {"state": state, "provider": provider}
    ARQUIVO_OAUTH_ESTADO.write_text(json.dumps(dados), encoding="utf-8")


def carregar_estado_oauth() -> Dict[str, str]:
    if not ARQUIVO_OAUTH_ESTADO.exists():
        return {}

    try:
        conteudo = ARQUIVO_OAUTH_ESTADO.read_text(encoding="utf-8")
        dados = json.loads(conteudo)
        if isinstance(dados, dict):
            return {str(chave): str(valor) for chave, valor in dados.items()}
    except Exception:
        return {}

    return {}


def limpar_estado_oauth() -> None:
    try:
        if ARQUIVO_OAUTH_ESTADO.exists():
            ARQUIVO_OAUTH_ESTADO.unlink()
    except Exception:
        pass


def carregar_usuarios_locais(config: Dict[str, object]) -> List[Dict[str, str]]:
    usuarios: List[Dict[str, str]] = []

    json_usuarios = config.get("LOCAL_USERS_JSON")
    if json_usuarios:
        try:
            dados = json.loads(str(json_usuarios))
            if isinstance(dados, list):
                for item in dados:
                    if isinstance(item, dict):
                        usuarios.append({str(k): str(v) for k, v in item.items()})
        except json.JSONDecodeError:
            pass

    email = str(config.get("LOCAL_LOGIN_EMAIL", "")).strip()
    senha = str(config.get("LOCAL_LOGIN_PASSWORD", "")).strip()
    nome = str(config.get("LOCAL_LOGIN_NAME", "Usuario local")).strip() or "Usuario local"
    if email and senha:
        usuarios.append({
            "email": email,
            "password": senha,
            "name": nome,
        })

    return usuarios


def autenticar_login_local(email: str, senha: str, usuarios: List[Dict[str, str]]) -> Dict[str, str] | None:
    email_normalizado = email.strip().lower()

    for usuario in usuarios:
        email_usuario = str(usuario.get("email", "")).strip().lower()
        senha_usuario = str(usuario.get("password", "")).strip()
        if email_usuario == email_normalizado and senha_usuario == senha:
            return {
                "provider": "local",
                "email": email_usuario,
                "name": str(usuario.get("name", email_usuario)).strip() or email_usuario,
            }

    return None


def obter_config_oauth(provider: str, config: Dict[str, object]) -> Dict[str, object] | None:
    if provider == "google":
        client_id = str(config.get("GOOGLE_CLIENT_ID", "")).strip()
        client_secret = str(config.get("GOOGLE_CLIENT_SECRET", "")).strip()
        redirect_uri = str(config.get("GOOGLE_REDIRECT_URI", "")).strip()
        if not client_id or not client_secret or not redirect_uri or "..." in redirect_uri:
            return None
        return {
            "provider": "google",
            "label": "Google",
            "client_id": client_id,
            "client_secret": client_secret,
            "redirect_uri": redirect_uri,
            "authorize_url": "https://accounts.google.com/o/oauth2/v2/auth",
            "token_url": "https://oauth2.googleapis.com/token",
            "userinfo_url": "https://openidconnect.googleapis.com/v1/userinfo",
            "scope": "openid email profile",
            "extra_params": {"prompt": "select_account"},
        }


    return None


def montar_url_autorizacao(provider: str, config: Dict[str, object]) -> str | None:
    oauth = obter_config_oauth(provider, config)
    if oauth is None:
        return None

    state = f"{provider}:{secrets.token_urlsafe(24)}"
    st.session_state["oauth_state"] = state
    salvar_estado_oauth(state, provider)

    params = {
        "client_id": oauth["client_id"],
        "redirect_uri": oauth["redirect_uri"],
        "response_type": "code",
        "scope": oauth["scope"],
        "state": state,
    }
    params.update(oauth.get("extra_params", {}))

    return f"{oauth['authorize_url']}?{urllib.parse.urlencode(params)}"


def criar_contexto_ssl() -> ssl.SSLContext:
    if certifi is not None:
        return ssl.create_default_context(cafile=certifi.where())
    return ssl.create_default_context()


def requisicao_json(url: str, data: Dict[str, str] | None = None, headers: Dict[str, str] | None = None) -> Dict[str, object]:
    headers = headers or {}
    contexto_ssl = criar_contexto_ssl()

    if data is None:
        request = urllib.request.Request(url, headers=headers)
    else:
        body = urllib.parse.urlencode(data).encode("utf-8")
        headers = {"Content-Type": "application/x-www-form-urlencoded", **headers}
        request = urllib.request.Request(url, data=body, headers=headers)

    with urllib.request.urlopen(request, timeout=20, context=contexto_ssl) as resposta:
        return json.loads(resposta.read().decode("utf-8"))


def trocar_code_por_token(provider: str, code: str, config: Dict[str, object]) -> Dict[str, object]:
    oauth = obter_config_oauth(provider, config)
    if oauth is None:
        raise ValueError(f"Configuracao de {provider} nao encontrada.")

    payload = {
        "code": code,
        "client_id": str(oauth["client_id"]),
        "client_secret": str(oauth["client_secret"]),
        "redirect_uri": str(oauth["redirect_uri"]),
        "grant_type": "authorization_code",
    }
    return requisicao_json(str(oauth["token_url"]), data=payload)


def decodificar_id_token_sem_validacao(id_token: str) -> Dict[str, object]:
    partes = id_token.split(".")
    if len(partes) < 2:
        return {}

    payload = partes[1]
    padding = "=" * (-len(payload) % 4)
    try:
        conteudo = urllib.parse.unquote_to_bytes(payload + padding)
    except Exception:
        conteudo = (payload + padding).encode("utf-8")

    try:
        import base64
        bruto = base64.urlsafe_b64decode(payload + padding)
        return json.loads(bruto.decode("utf-8"))
    except Exception:
        try:
            return json.loads(conteudo.decode("utf-8"))
        except Exception:
            return {}


def obter_usuario_oauth(provider: str, token: Dict[str, object], config: Dict[str, object]) -> Dict[str, str]:
    oauth = obter_config_oauth(provider, config)
    if oauth is None:
        raise ValueError(f"Configuracao de {provider} nao encontrada.")

    access_token = str(token.get("access_token", "")).strip()
    if access_token:
        dados = requisicao_json(
            str(oauth["userinfo_url"]),
            headers={"Authorization": f"Bearer {access_token}"},
        )
        return {
            "provider": provider,
            "email": str(dados.get("email") or dados.get("preferred_username") or "").strip(),
            "name": str(dados.get("name") or dados.get("given_name") or dados.get("preferred_username") or "Usuario").strip(),
        }

    id_token = str(token.get("id_token", "")).strip()
    dados = decodificar_id_token_sem_validacao(id_token)
    return {
        "provider": provider,
        "email": str(dados.get("email") or dados.get("preferred_username") or "").strip(),
        "name": str(dados.get("name") or dados.get("given_name") or dados.get("preferred_username") or "Usuario").strip(),
    }


def processar_callback_oauth(config: Dict[str, object]) -> None:
    params = obter_query_params()
    code = params.get("code", "")
    state = params.get("state", "")
    erro = params.get("error", "")

    if erro:
        st.error(f"Falha no login SSO: {erro}.")
        limpar_query_params()
        return

    if not code or not state:
        return

    estado_salvo = carregar_estado_oauth()
    state_esperado = st.session_state.get("oauth_state", "") or estado_salvo.get("state", "")
    if not state_esperado or state != state_esperado:
        st.error("Nao foi possivel validar a resposta do provedor de login.")
        limpar_query_params()
        limpar_estado_oauth()
        return

    provider = estado_salvo.get("provider", "") or state.split(":", 1)[0]

    try:
        token = trocar_code_por_token(provider, code, config)
        usuario = obter_usuario_oauth(provider, token, config)
    except Exception as exc:
        st.error(f"Falha ao concluir o login com {provider.title()}: {exc}")
        limpar_query_params()
        limpar_estado_oauth()
        return

    if not usuario.get("email"):
        st.error("O provedor retornou o login, mas sem um e-mail utilizavel.")
        limpar_query_params()
        limpar_estado_oauth()
        return

    st.session_state["auth_user"] = usuario
    limpar_query_params()
    limpar_estado_oauth()
    st.rerun()


def usuario_autenticado() -> bool:
    return bool(st.session_state.get("auth_user"))


def sair() -> None:
    st.session_state.pop("auth_user", None)
    st.session_state.pop("oauth_state", None)
    limpar_query_params()
    limpar_estado_oauth()
    st.rerun()


def renderizar_botao_oauth(provider: str, config: Dict[str, object], compact: bool = False) -> None:
    oauth = obter_config_oauth(provider, config)
    if oauth is None:
        st.warning(f"Login com {provider.title()} ainda nao esta configurado no secrets.toml.")
        return

    url = montar_url_autorizacao(provider, config)
    if not url:
        st.warning(f"Nao foi possivel montar a URL de login do {provider.title()}.")
        return

    if compact:
        google_icon = "https://www.gstatic.com/firebasejs/ui/2.0.0/images/auth/google.svg"
        card_html = f"""
        <a href="{url}" target="_self" style="text-decoration:none;">
            <div style="border:1px solid #374151;border-radius:16px;padding:22px 18px;display:flex;flex-direction:column;align-items:center;justify-content:center;gap:14px;min-height:210px;background:#111827;">
                <img src="{google_icon}" alt="Google" style="width:42px;height:42px;" />
                <div style="color:#f9fafb;font-size:1.05rem;font-weight:600;">Entrar com Google</div>
                <div style="color:#9ca3af;font-size:0.9rem;text-align:center;">Use sua conta Google para acessar o sistema.</div>
            </div>
        </a>
        """
        st.markdown(card_html, unsafe_allow_html=True)
    else:
        label = f"Entrar com {oauth['label']}"
        if hasattr(st, "link_button"):
            st.link_button(label, url, use_container_width=True)
        else:
            st.markdown(f"[{label}]({url})")
        st.caption(f"Redirect URI configurada: {oauth['redirect_uri']}")


def renderizar_tela_login(config: Dict[str, object]) -> None:
    _render_notranslate_guard()
    st.markdown(
        """
        <style>
        .login-shell {max-width: 1180px; margin: 0 auto; padding-top: 1.5rem;}
        .login-title {text-align:center; margin-bottom: 0.35rem;}
        .login-subtitle {text-align:center; color:#9ca3af; margin-bottom: 1.8rem;}
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.markdown('<div class="login-shell">', unsafe_allow_html=True)
    st.markdown('<h1 class="login-title">Bem-vindo ao Projeto TCC</h1>', unsafe_allow_html=True)
    st.markdown('<div class="login-subtitle">Entre com login local ou com sua conta Google.</div>', unsafe_allow_html=True)

    col_local, col_google = st.columns([1.25, 0.9], gap='large')

    with col_local:
        with st.container(border=True):
            st.markdown('### Login com e-mail e senha')
            usuarios = carregar_usuarios_locais(config)
            with st.form('login_local'):
                email = st.text_input('E-mail')
                senha = st.text_input('Senha', type='password')
                enviar = st.form_submit_button('Entrar', use_container_width=True)
            if enviar:
                usuario = autenticar_login_local(email, senha, usuarios)
                if usuario is None:
                    st.error('E-mail ou senha invalidos.')
                else:
                    st.session_state['auth_user'] = usuario
                    st.rerun()
            if not usuarios:
                st.info('Configure LOCAL_LOGIN_EMAIL e LOCAL_LOGIN_PASSWORD no secrets.toml para habilitar o login local.')

    with col_google:
        with st.container(border=True):
            st.markdown('### Google')
            renderizar_botao_oauth('google', config, compact=True)

    st.markdown('</div>', unsafe_allow_html=True)
    st.stop()

def renderizar_aplicacao_principal() -> None:
    _render_notranslate_guard()
    usuario = st.session_state.get("auth_user", {})

    with st.sidebar:
        st.success(f"Conectado como: {usuario.get('name', 'Usuario')}")
        st.caption(f"E-mail: {usuario.get('email', '-')}")
        st.caption(f"Acesso: {str(usuario.get('provider', 'local')).title()}")
        if st.button("Sair", use_container_width=True):
            sair()

    st.title("Sistema de Dimensionamento Eletrico Residencial")
    st.caption(
        "Versao corrigida para previsao minima de cargas pela NBR 5410, resumo simplificado de demanda CPFL/GED-13 "
        "e caracteristicas do padrao de entrada conforme a tabela 1A usada nas aulas."
    )

    with st.sidebar:
        st.header("Projeto")
        nome_projeto = st.text_input("Nome do projeto", value="Meu Projeto")
        responsavel = st.text_input("Responsavel", value="")
        origem_comodos = st.radio(
            "Origem dos comodos",
            options=["DXF", "Manual"],
            index=0,
            help="Em DXF, os comodos sao desenhados sobre a planta e o formulario herda nome, area e perimetro automaticamente.",
        )
        if origem_comodos == "Manual":
            numero_comodos = st.number_input("Quantidade de comodos", min_value=1, max_value=30, value=5, step=1)
        else:
            numero_comodos = 0
            st.caption("A quantidade de comodos sera definida pela selecao feita no DXF.")

        st.header("Padrao de Entrada")
        fase_padrao = st.selectbox(
            "Fase do padrao de entrada",
            options=["Automatico", "Monofasico", "Bifasico", "Trifasico"],
            index=0,
            help="Se ficar em Automatico, o app sugere a fase com base na categoria. Se voce quiser, pode fixar manualmente.",
        )

    comodos_importados = []
    if origem_comodos == "DXF":
        comodos_importados = renderizar_importacao_dxf()
        numero_comodos = len(comodos_importados)
        if numero_comodos == 0:
            st.warning("Desenhe e salve ao menos um comodo no DXF para continuar com o formulario abaixo.")

    st.subheader("2) Cadastro dos comodos" if origem_comodos == "DXF" else "1) Cadastro dos comodos")

    comodos = []
    equipamentos_gerais_demanda = []

    for i in range(int(numero_comodos)):
        comodo_importado = comodos_importados[i] if i < len(comodos_importados) else None
        with st.expander(f"Comodo {i + 1}", expanded=(i == 0)):
            col1, col2, col3 = st.columns(3)
            with col1:
                nome_padrao = comodo_importado["nome"] if comodo_importado else f"Comodo {i + 1}"
                nome_comodo = st.text_input(
                    f"Nome do comodo {i + 1}",
                    key=f"nome_{i}",
                    value=nome_padrao,
                    disabled=bool(comodo_importado),
                )
            with col2:
                tipos_comodo_opcoes = [
                    "quarto",
                    "suite",
                    "sala",
                    "cozinha",
                    "banheiro",
                    "area_servico",
                    "lavanderia",
                    "copa",
                    "copa_cozinha",
                    "escritorio",
                    "circulacao",
                    "corredor",
                    "hall",
                    "lavabo",
                    "closet",
                    "varanda",
                    "garagem",
                    "sotao",
                    "subsolo",
                    "casa_maquinas",
                    "sala_bombas",
                    "barrilete",
                    "outro",
                ]
                tipos_comodo_labels = {
                    "quarto": "Quarto",
                    "suite": "Suite",
                    "sala": "Sala",
                    "cozinha": "Cozinha",
                    "banheiro": "Banheiro",
                    "area_servico": "Area de Servico",
                    "lavanderia": "Lavanderia",
                    "copa": "Copa",
                    "copa_cozinha": "Copa/Cozinha",
                    "escritorio": "Escritorio",
                    "circulacao": "Circulacao",
                    "corredor": "Corredor",
                    "hall": "Hall",
                    "lavabo": "Lavabo",
                    "closet": "Closet",
                    "varanda": "Varanda",
                    "garagem": "Garagem",
                    "sotao": "Sotao",
                    "subsolo": "Subsolo",
                    "casa_maquinas": "Casa de Maquinas",
                    "sala_bombas": "Sala de Bombas",
                    "barrilete": "Barrilete",
                    "outro": "Outro",
                }
                tipo_padrao = comodo_importado.get("tipo_sugerido", _infer_tipo_comodo(nome_padrao)) if comodo_importado else "quarto"
                if tipo_padrao not in tipos_comodo_opcoes:
                    tipo_padrao = "outro"
                tipo_comodo = st.selectbox(
                    f"Tipo do comodo {i + 1}",
                    options=tipos_comodo_opcoes,
                    index=tipos_comodo_opcoes.index(tipo_padrao),
                    format_func=lambda x: tipos_comodo_labels.get(x, x),
                    key=f"tipo_{i}",
                )
            with col3:
                area = st.number_input(
                    f"Area (m2) - {i + 1}",
                    min_value=0.01,
                    value=float(comodo_importado["area"]) if comodo_importado else 10.0,
                    step=0.1,
                    key=f"area_{i}",
                    disabled=bool(comodo_importado),
                )

            col4, col5 = st.columns(2)
            with col4:
                perimetro = st.number_input(
                    f"Perimetro (m) - {i + 1}",
                    min_value=0.01,
                    value=float(comodo_importado["perimetro"]) if comodo_importado else 12.0,
                    step=0.1,
                    key=f"per_{i}",
                    disabled=bool(comodo_importado),
                )
            with col5:
                bancadas_validas = 0
                if normalizar_ambiente(tipo_comodo) in {"cozinha", "copa", "copa_cozinha", "area_servico", "lavanderia"}:
                    bancadas_validas = st.number_input(
                        f"Qtde. de bancadas >= 0,30 m - {i + 1}",
                        min_value=0,
                        value=0,
                        step=1,
                        key=f"bancadas_{i}",
                    )
                else:
                    st.markdown("**Bancadas:** nao se aplica")

            st.markdown("**TUEs do comodo**")
            qtd_tues = st.number_input(
                f"Quantidade de TUEs - {i + 1}",
                min_value=0,
                max_value=10,
                value=0,
                step=1,
                key=f"qtd_tue_{i}",
            )
            tues = []

            for j in range(int(qtd_tues)):
                c1, c2, c3 = st.columns([2, 1, 2])
                with c1:
                    nome_eq = st.text_input(f"Equipamento {j + 1}", key=f"eq_nome_{i}_{j}", value=f"Equipamento {j + 1}")
                with c2:
                    potencia_w = st.number_input(
                        f"Potencia (W) {j + 1}",
                        min_value=0.0,
                        value=1000.0,
                        step=100.0,
                        key=f"eq_pot_{i}_{j}",
                    )
                with c3:
                    categoria_demanda = st.selectbox(
                        f"Categoria demanda {j + 1}",
                        options=list(CATEGORIAS_DEMANDA.keys()),
                        format_func=lambda x: f"{x}) {CATEGORIAS_DEMANDA[x]}",
                        key=f"eq_cat_{i}_{j}",
                    )

                registro_eq = {
                    "nome": nome_eq,
                    "potencia_w": potencia_w,
                    "categoria_demanda": categoria_demanda,
                    "comodo": nome_comodo,
                }
                tues.append(registro_eq)
                equipamentos_gerais_demanda.append(registro_eq)

            comodos.append(
                {
                    "nome": nome_comodo,
                    "tipo": tipo_comodo,
                    "area": area,
                    "perimetro": perimetro,
                    "bancadas_validas": bancadas_validas,
                    "tues": tues,
                    "vertices": (comodo_importado.get("vertices") if comodo_importado else []),
                    "centroide_x": (comodo_importado.get("centroide_x") if comodo_importado else None),
                    "centroide_y": (comodo_importado.get("centroide_y") if comodo_importado else None),
                }
            )

    calcular_projeto = st.button("Calcular projeto", type="primary")
    if calcular_projeto:
        projeto_calculado = montar_projeto_calculado(
            nome_projeto=nome_projeto,
            responsavel=responsavel,
            fase_padrao=fase_padrao,
            comodos=comodos,
            equipamentos_gerais_demanda=equipamentos_gerais_demanda,
            segments=st.session_state.get(_session_key('segments')) if origem_comodos == 'DXF' else [],
            endpoints=st.session_state.get(_session_key('endpoints')) if origem_comodos == 'DXF' else [],
            comodos_importados=comodos_importados if origem_comodos == 'DXF' else [],
        )
        st.session_state['ultimo_projeto_calculado'] = projeto_calculado
        st.session_state['mostrar_dimensionamento'] = False

    projeto_calculado = st.session_state.get('ultimo_projeto_calculado')
    if projeto_calculado:
        st.subheader("2) Tabela de cargas por comodo")
        st.dataframe(projeto_calculado['df_resultados'], use_container_width=True)

        st.subheader("3) Totais instalados")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Iluminacao total", f"{formatar_numero_br(projeto_calculado['total_ilum_va'], 1)} VA")
        c2.metric("TUG total", f"{formatar_numero_br(projeto_calculado['total_tug_va'], 1)} VA")
        c3.metric("TUE total", f"{formatar_numero_br(projeto_calculado['total_tue_w'], 1)} W")
        c4.metric("Carga instalada total", f"{formatar_numero_br(projeto_calculado['potencia_instalada_total_w'], 1)} W")

        st.info(
            "Observacao: a NBR 5410 define a carga minima de iluminacao em VA para dimensionamento; "
            "isso nao determina automaticamente a quantidade real de lampadas/luminarias do projeto."
        )

        st.subheader("4) Resumo simplificado de demanda CPFL / GED-13")
        st.dataframe(projeto_calculado['df_demanda'], use_container_width=True)
        st.metric("Demanda total simplificada", f"{formatar_numero_br(projeto_calculado['total_demanda_w'], 1)} W")

        st.subheader("5) Caracter?sticas do Padr?o de Entrada")
        _render_df_html_table(projeto_calculado['df_padrao'])
        st.caption(
            "Categoria sugerida conforme a tabela 1A da GED-13 usada nas aulas. "
            "Se necessario, a fase pode ser ajustada manualmente no menu lateral."
        )

        st.subheader("6) Exportacao")
        st.caption("O XLSX passa a incluir a aba de dimensionamento assim que essa etapa for concluida.")

        abrir_dimensionamento = st.button(
            'Dimensionamento',
            key='abrir_dimensionamento',
            use_container_width=True,
            disabled=not bool(projeto_calculado.get('segments')),
        )
        if abrir_dimensionamento:
            st.session_state['mostrar_dimensionamento'] = True

        df_dimensionamento = None
        df_pontos_dimensionamento = None
        if st.session_state.get('mostrar_dimensionamento'):
            df_dimensionamento, df_pontos_dimensionamento = renderizar_dimensionamento(projeto_calculado)

        excel_bytes = gerar_excel_bytes(
            df_resultados=projeto_calculado['df_resultados'],
            df_demanda=projeto_calculado['df_demanda'],
            df_padrao=projeto_calculado['df_padrao'],
            df_resumo=projeto_calculado['df_resumo'],
            df_dimensionamento=df_dimensionamento,
            df_pontos_dimensionamento=df_pontos_dimensionamento,
        )

        st.download_button(
            label="Baixar relatorio em XLSX",
            data=excel_bytes,
            file_name=f"relatorio_eletrico_{projeto_calculado['nome_projeto'].replace(' ', '_').lower()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.subheader("7) Conferencia rapida")
        st.markdown(
            f"""
            **Projeto:** {projeto_calculado['nome_projeto']}  
            **Responsavel:** {projeto_calculado['responsavel'] or '-'}  
            **Quantidade de comodos:** {len(projeto_calculado['comodos'])}  
            **Soma iluminacao:** {formatar_numero_br(projeto_calculado['total_ilum_va'], 1)} VA  
            **Soma TUG:** {formatar_numero_br(projeto_calculado['total_tug_va'], 1)} VA  
            **Soma TUE:** {formatar_numero_br(projeto_calculado['total_tue_w'], 1)} W  
            **Carga instalada total:** {formatar_numero_br(projeto_calculado['potencia_instalada_total_w'], 1)} W  
            **Demanda simplificada:** {formatar_numero_br(projeto_calculado['total_demanda_w'], 1)} W  
            **Padrao sugerido:** {projeto_calculado['padrao_entrada']['Categoria']} / {projeto_calculado['padrao_entrada']['Fase']} / Caixa {projeto_calculado['padrao_entrada']['Tipo de Caixa']} / Disj. {projeto_calculado['padrao_entrada']['Disjuntor']} A / Eletroduto {projeto_calculado['padrao_entrada']['Medida do Eletroduto']}
            """
        )



def main() -> None:
    config = carregar_config_autenticacao()
    processar_callback_oauth(config)

    if not usuario_autenticado():
        renderizar_tela_login(config)

    renderizar_aplicacao_principal()


main()
